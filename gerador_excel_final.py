"""
Gerador de Excel — formato CM5 exato
Reproduz as 8 abas do arquivo Escala_ClinicaMedica_5ANO_Grupo_A_T6_vJUNIA
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import io
import json

# ── Paleta de cores exata do CM5 ─────────────────────────────────────────────
C = {
    # Headers
    "H1":  "1F4E79",  # azul escuro (título)
    "H2":  "2E75B6",  # azul médio (sub-header)
    "H3":  "D6E4F0",  # azul claro (header datas)
    "FDS": "7FD1C4",  # TEAL (fim de semana) — colorido e visual, letras pretas
    "TER": "FFB3B3",  # rosa mais forte (terça ★16h) — antes era apagado
    "ENAMED": "FF9E9E",  # rosa/vermelho ENAMED (quinta) — mais visível
    # SGs - cor da linha de nome
    "SG1": "D6E4F7", "SG2": "D4EED4", "SG3": "FFF3CD",
    "SG4": "FFD6D6", "SG5": "EDD9F5", "SG6": "D9F5E8",
    "SG7": "FDE8D8", "SG8": "E8F8E8",
    # Locais - cor das células de dados
    "ENF": "AED6F1",  # azul claro (enfermaria)
    "PS":  "A9DFBF",  # verde claro (PS)
    "AMB": "FAD7A0",  # laranja claro (ambulatório)
    "LOC4":"D7BDE2",  # lilás
    "LOC5":"F9E79F",  # amarelo
    "LOC6":"FADBD8",  # rosa
    # Especiais
    "FDS_CELL": "D6F0EA",  # fim de semana sem atividade (teal bem claro)
    "FDS_PLT":  "D7BDE2",  # plantão FDS / PA★
    "TARDE_R":  "FFF2CC",  # tarde reduzida
    "VERDE":    "E2EFDA",  # área verde
    "TOTAL":    "FFF2CC",  # total horas
    "CIND":     "FFFFFF",
    "ENF_FDS_H":"EBF5FB",  # header plantões enf fds
    "PS_FDS_H": "EAFAF1",  # header plantões ps fds
    # Alternância linhas
    "ALT1": "EEF3FB",
    "ALT2": "FFFFFF",
    "SEP":  "2E75B6",  # separador de SG
}

DIAS_PT = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]

def _normalizar_data(data_str, ano="2026"):
    """Normaliza data para formato DD/MM/YYYY."""
    if not data_str: return ""
    s = str(data_str).strip()
    # Já está completo
    if len(s) == 10 and "/" in s: return s  # DD/MM/YYYY
    if len(s) == 10 and "-" in s:  # YYYY-MM-DD
        parts = s.split("-")
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    # DD/MM — adicionar ano
    if len(s) == 5 and "/" in s:
        return f"{s}/{ano}"
    if len(s) == 5 and "-" in s:
        parts = s.split("-")
        return f"{parts[1]}/{parts[0]}/{ano}"
    return s



def _cor(rgb):
    return PatternFill("solid", fgColor=rgb)

def _borda(color="BFBFBF", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _cel(ws, row, col, val="", bold=False, bg=None, fc="000000",
         halign="center", valign="center", wrap=False, sz=9, italic=False,
         border=True, border_color="BFBFBF"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fc, size=sz, italic=italic)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if border:
        c.border = _borda(border_color)
    if bg:
        c.fill = _cor(bg)
    return c

def _header(ws, row, col, val, bg=C["H1"], fc="FFFFFF", sz=11, bold=True, span=None):
    c = _cel(ws, row, col, val, bold=bold, bg=bg, fc=fc, sz=sz)
    if span and span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 22
    return c

def _cor_local(local_nome, locais_cfg):
    """Retorna cor da célula baseada no local."""
    cores = [C["ENF"], C["PS"], C["AMB"], C["LOC4"], C["LOC5"], C["LOC6"]]
    nomes = [l.get("nome","").lower() for l in locais_cfg]
    nome_lower = local_nome.lower()
    for i, n in enumerate(nomes):
        if n and (n in nome_lower or nome_lower in n):
            return cores[i % len(cores)]
    return C["ENF"]

def _cor_sg(sg_num):
    key = f"SG{sg_num}"
    return C.get(key, C["SG1"])

def _gerar_datas(data_inicio_str, num_semanas):
    """Retorna lista de semanas, cada uma com 7 datas."""
    try:
        d0 = date.fromisoformat(data_inicio_str)
    except:
        d0 = date.today()
    return [[d0 + timedelta(weeks=s, days=d) for d in range(7)] for s in range(num_semanas)]

def _turno_compacto(entrada):
    """Converte entrada da IA para formato compacto: Enf(M+T)"""
    if not entrada or entrada in ("—", "-", "", "None"):
        return "—"
    return str(entrada)

def gerar_excel_completo(dados, config):
    """
    dados: dict com calendario_rodizio, escala_detalhada, resumo_horas, auditoria
    config: dict com especialidade, grupo, turma, data_inicio, num_semanas, locais, alunos_por_sg
    """
    esp = config.get("especialidade","")
    grupo = config.get("grupo","")
    turma = config.get("turma","")
    ano = config.get("ano_curso","")
    titulo = f"{esp} — {ano} — {grupo} / {turma}"
    
    locais_cfg = config.get("locais", [])
    alunos_por_sg = config.get("alunos_por_sg", {})
    num_semanas = int(config.get("num_semanas", 8))
    data_inicio = config.get("data_inicio", str(date.today()))
    rodizio_desc = config.get("rodizio_desc", "")
    regras = config.get("regras_especiais", {})
    
    semanas = _gerar_datas(data_inicio, num_semanas)
    
    # Detectar ano da escala
    ano_escala = "2026"
    try:
        ano_escala = date.fromisoformat(data_inicio).strftime("%Y")
    except: pass

    # Montar índice de escala detalhada por aluno/data
    escala_det = dados.get("escala_detalhada", [])
    # {nome: {data_normalizada: "turno_compacto"}}
    escala_por_aluno = {}
    for entry in escala_det:
        alunos = entry.get("alunos", [])
        if isinstance(entry.get("nome"), str) and entry.get("nome"):
            alunos = [entry["nome"]]
        data_raw = str(entry.get("data",""))
        data_str = _normalizar_data(data_raw, ano_escala)
        turno = entry.get("turno","")
        local = entry.get("local","")
        abrev = _abrev_local(local, locais_cfg)
        codigo = _codigo_turno(turno, local, locais_cfg)
        valor = f"{abrev}({codigo})" if codigo else f"{abrev}"
        if entry.get("plantao"):
            valor = "★" + valor   # plantão de complemento (em outro serviço)
        for nome in alunos:
            if nome not in escala_por_aluno:
                escala_por_aluno[nome] = {}
            if data_str in escala_por_aluno[nome]:
                existing = escala_por_aluno[nome][data_str]
                if valor not in existing:
                    escala_por_aluno[nome][data_str] = _combinar(existing, valor)
            else:
                escala_por_aluno[nome][data_str] = valor

    # Índice por PERÍODO (M/T/C) — usado pela Escala Individual para mostrar a Área Verde
    # explicitamente no período de dia útil em que o aluno está livre.
    def _bucket_periodo(turno):
        t = (turno or "").lower()
        if "cind" in t or "noit" in t or "noturn" in t:
            return "C"
        if "tard" in t or "reduz" in t:
            return "T"
        return "M"
    escala_por_aluno_per = {}   # nome -> {data_str -> {"M":info, "T":info, "C":info}}
    for entry in escala_det:
        alunos = entry.get("alunos", [])
        if isinstance(entry.get("nome"), str) and entry.get("nome"):
            alunos = [entry["nome"]]
        data_str = _normalizar_data(str(entry.get("data","")), ano_escala)
        turno = entry.get("turno","")
        local = entry.get("local","")
        b = _bucket_periodo(turno)
        info = {"abrev": _abrev_local(local, locais_cfg), "local": local,
                "plantao": bool(entry.get("plantao")),
                "reduz": ("reduz" in turno.lower() or "12-16" in str(entry.get("horario","")))}
        for nome in alunos:
            escala_por_aluno_per.setdefault(nome, {}).setdefault(data_str, {})[b] = info

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _aba_resumo_geral(wb, titulo, config, dados, semanas)
    _aba_alunos(wb, titulo, alunos_por_sg, config)
    _aba_calendario_rodizio(wb, titulo, dados.get("calendario_rodizio",[]), config, semanas)
    _aba_escala_nominal(wb, titulo, escala_det, config, semanas, locais_cfg, ano_escala)
    _aba_resumo_horas(wb, titulo, dados.get("resumo_horas",[]), config, semanas, locais_cfg, dados)
    _aba_regras(wb, titulo, config, dados)
    _aba_escala_subgrupo(wb, titulo, alunos_por_sg, escala_por_aluno, config, semanas, locais_cfg)
    _aba_escala_individual(wb, titulo, alunos_por_sg, escala_por_aluno_per, config, semanas, locais_cfg)
    _aba_por_servico(wb, titulo, escala_det, config, semanas, locais_cfg, ano_escala)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _abrev_local(local_nome, locais_cfg):
    """Retorna abreviação do local."""
    for l in locais_cfg:
        n = l.get("nome","")
        if n and (n.lower() in local_nome.lower() or local_nome.lower() in n.lower()):
            ab = l.get("abrev","")
            return ab if ab else n[:3]
    # Fallback: primeiras 3 letras
    return local_nome[:3] if local_nome else "?"

def _codigo_turno(turno, local, locais_cfg):
    """Converte turno para código compacto."""
    t = turno.lower() if turno else ""
    if "fds" in t or "★" in turno or "pa" in t.lower():
        return "FDS"
    if "cinderela" in t or t == "c": return "C"
    if "reduz" in t or t == "r": return "R"
    if "manhã" in t or t == "m" or "manha" in t: return "M"
    if "tarde" in t or t == "t": return "T"
    if "noite" in t or t == "n": return "N"
    if "+" in t: return t.upper()
    return t[:1].upper() if t else "M"

def _combinar(v1, v2):
    """Combina dois valores de turno."""
    if v1 == v2: return v1
    # Extrair local e turno
    def parse(v):
        if "(" in v and ")" in v:
            loc = v[:v.index("(")]
            t = v[v.index("(")+1:v.index(")")]
            return loc, t
        return v, ""
    loc1, t1 = parse(v1)
    loc2, t2 = parse(v2)
    if loc1 == loc2:
        # Mesmo local, combinar turnos
        turnos = set()
        for t in [t1, t2]:
            for part in t.split("+"):
                turnos.add(part)
        ordem = ["M","T","R","C","F","N","FDS"]
        combined = "+".join([x for x in ordem if x in turnos])
        return f"{loc1}({combined})" if combined else loc1
    return f"{v1}/{v2}"


# ── ABA 1: RESUMO GERAL ───────────────────────────────────────────────────────
def _aba_resumo_geral(wb, titulo, config, dados, semanas):
    ws = wb.create_sheet("Resumo Geral")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55

    esp = config.get("especialidade","")
    grupo = config.get("grupo","")
    turma = config.get("turma","")
    ano = config.get("ano_curso","")
    locais_cfg = config.get("locais",[])
    alunos_por_sg = config.get("alunos_por_sg",{})
    regras = config.get("regras_especiais",{})
    rodizio = config.get("rodizio_desc","")

    d_ini = semanas[0][0].strftime("%d/%m/%Y") if semanas else ""
    d_fim = semanas[-1][-1].strftime("%d/%m/%Y") if semanas else ""
    n_sem = len(semanas)
    n_alunos = sum(len(v) for v in alunos_por_sg.values())

    _header(ws, 1, 1, f"{esp} — {ano} — {grupo} / {turma}", span=2, sz=13)
    _header(ws, 2, 1, f"Escala de Estágio  |  {d_ini} – {d_fim}  |  {n_sem} semanas  |  Rodízio {rodizio[:50]}", span=2, bg=C["H2"], sz=10)
    ws.row_dimensions[2].height = 16

    dados_resumo = [
        ("Período", f"{d_ini} a {d_fim} ({n_sem} semanas)"),
        ("Grupo / Turma", f"{grupo} — {turma} — {ano}"),
        ("Total de alunos", str(n_alunos)),
        ("Subgrupos", " | ".join([f"SG{k}({len(v)}al)" for k,v in sorted(alunos_por_sg.items(), key=lambda x:int(x[0]))])),
        ("Locais", " · ".join([l.get("nome","") for l in locais_cfg])),
        ("Rodízio", rodizio),
    ]
    # Regras de cada local
    for l in locais_cfg:
        linha_m = l.get("manha","")
        linha_t = l.get("tarde","")
        dados_resumo.append((l.get("nome",""), f"Manhã: {linha_m} | Tarde: {linha_t}"))
    dados_resumo += [
        ("Quinta-feira", regras.get("quinta","ENAMED — sem campo tarde")),
        ("Terça-feira", regras.get("terca","Aula — campo tarde encurtado")),
        ("Limite CH", f"{regras.get('limite_ch',40)}h/sem | Absoluto: {regras.get('limite_abs',43)}h"),
        ("FDS", regras.get("fds","")),
    ]

    row = 4
    for label, val in dados_resumo:
        if not val: continue
        _cel(ws, row, 1, label, bold=True, bg=C["H3"], halign="left", sz=9)
        _cel(ws, row, 2, val, halign="left", sz=9, wrap=True)
        ws.row_dimensions[row].height = 15
        row += 1

    # Legenda
    row += 1
    _header(ws, row, 1, "LEGENDA DE LOCAIS", span=2, bg=C["H2"])
    row += 1
    cores_loc = [C["ENF"], C["PS"], C["AMB"], C["LOC4"], C["LOC5"], C["LOC6"]]
    for i, l in enumerate(locais_cfg):
        cor = cores_loc[i % len(cores_loc)]
        ab = l.get("abrev", l.get("nome","")[:3])
        _cel(ws, row, 1, ab, bold=True, bg=cor, sz=9)
        _cel(ws, row, 2, l.get("nome",""), halign="left", bg=cor, sz=9)
        row += 1

    # Legenda de turnos
    row += 1
    _header(ws, row, 1, "LEGENDA DE TURNOS", span=2, bg=C["H2"])
    row += 1
    turnos_leg = [
        ("M", "Manhã (horário conforme local)", None),
        ("T", "Tarde completa", None),
        ("R", "Tarde reduzida (fundo amarelo)", C["TARDE_R"]),
        ("F", "Período FDS manhã Enf/PS", C["FDS"]),
        ("★", "Plantão de complemento em outro serviço — fundo lilás", C["FDS_PLT"]),
        ("AV", "Área Verde — tarde livre", C["VERDE"]),
        ("C", "Cinderela", None),
        ("—", "Sem atividade (FDS sem período)", C["FDS_CELL"]),
    ]
    for ab, desc, cor in turnos_leg:
        _cel(ws, row, 1, ab, bold=True, bg=cor or "FFFFFF", sz=9)
        _cel(ws, row, 2, desc, halign="left", bg=cor or "FFFFFF", sz=9)
        row += 1

    # Auditoria
    audit = dados.get("auditoria",{})
    if audit:
        row += 1
        _header(ws, row, 1, "AUDITORIA", span=2, bg=C["H1"])
        row += 1
        status = "✅ APROVADA" if audit.get("aprovado") else "❌ COM ERROS"
        cor_st = "C6EFCE" if audit.get("aprovado") else "FFC7CE"
        _cel(ws, row, 1, "Status", bold=True, bg="F2F2F2", sz=9)
        _cel(ws, row, 2, status, bold=True, bg=cor_st, sz=9, halign="left")
        for err in audit.get("erros",[]):
            row += 1
            _cel(ws, row, 1, "ERRO", bold=True, bg="FFC7CE", sz=9)
            _cel(ws, row, 2, err, bg="FFC7CE", sz=9, halign="left", wrap=True)
        for av in audit.get("avisos",[]):
            row += 1
            _cel(ws, row, 1, "AVISO", bold=True, bg="FFEB9C", sz=9)
            _cel(ws, row, 2, av, bg="FFEB9C", sz=9, halign="left", wrap=True)


# ── ABA 2: ALUNOS ────────────────────────────────────────────────────────────
def _aba_alunos(wb, titulo, alunos_por_sg, config):
    ws = wb.create_sheet("Alunos")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 10

    turma = config.get("turma","")
    grupo = config.get("grupo","")
    ano = config.get("ano_curso","")

    _header(ws, 1, 1, f"LISTA DE ALUNOS — {grupo} / {turma} — {ano}", span=6, sz=11)
    for i, h in enumerate(["Nº","SG","Nome Completo","RA","Turma","Par"],1):
        _cel(ws, 2, i, h, bold=True, bg=C["H2"], fc="FFFFFF", sz=9)

    ra_map = config.get("ra_por_aluno", {}) or {}
    row = 3; num = 1
    for sg_key in sorted(alunos_por_sg.keys(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        sg_num = int(sg_key) if sg_key.isdigit() else 1
        cor = _cor_sg(sg_num)
        for i_al, nome in enumerate(alunos_por_sg[sg_key]):
            cor_row = C["ALT1"] if i_al % 2 == 0 else C["ALT2"]
            _cel(ws, row, 1, num, bg=cor_row, sz=9)
            _cel(ws, row, 2, f"SG{sg_num}", bold=True, bg=cor, sz=9)
            _cel(ws, row, 3, nome, halign="left", bg=cor_row, sz=9)
            _cel(ws, row, 4, str(ra_map.get(str(nome).strip(), "")), bg=cor_row, sz=9)  # RA
            _cel(ws, row, 5, turma, bg=cor_row, sz=9)
            _cel(ws, row, 6, "", bg=cor_row, sz=9)  # Par
            num += 1; row += 1


# ── ABA 3: CALENDÁRIO DE RODÍZIO ─────────────────────────────────────────────
def _aba_calendario_rodizio(wb, titulo, calendario, config, semanas):
    ws = wb.create_sheet("Calendário de Rodízio")
    ws.sheet_view.showGridLines = False

    grupo = config.get("grupo","")
    turma = config.get("turma","")
    rodizio_desc = config.get("rodizio_desc","")
    locais_cfg = config.get("locais",[])
    alunos_por_sg = config.get("alunos_por_sg",{})

    sgs = sorted(alunos_por_sg.keys(), key=lambda x: int(x[0]) if x[0].isdigit() else 0)
    n_sgs = len(sgs)

    _header(ws, 1, 1, f"CALENDÁRIO DE RODÍZIO — {grupo} / {turma}", span=2+n_sgs, sz=11)
    _header(ws, 2, 1, rodizio_desc[:120], span=2+n_sgs, bg=C["H2"], sz=9)

    # Headers SGs
    _cel(ws, 3, 1, "Par / SG", bold=True, bg=C["H2"], fc="FFFFFF", sz=9)
    _cel(ws, 3, 2, "Período", bold=True, bg=C["H2"], fc="FFFFFF", sz=9)
    for i, sg in enumerate(sgs):
        n_al = len(alunos_por_sg.get(sg,[]))
        _cel(ws, 3, 3+i, f"SG{sg}({n_al})", bold=True, bg=C["H2"], fc="FFFFFF", sz=9)

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 22
    for i in range(n_sgs):
        ws.column_dimensions[get_column_letter(3+i)].width = 22

    cores_loc = [C["ENF"], C["PS"], C["AMB"], C["LOC4"], C["LOC5"], C["LOC6"]]

    def _clarear(hexc, fator):
        """Clareia uma cor hex em direção ao branco (fator 0=igual, 1=branco)."""
        try:
            r = int(hexc[0:2], 16); g = int(hexc[2:4], 16); b = int(hexc[4:6], 16)
        except Exception:
            return hexc
        r = int(r + (255 - r) * fator); g = int(g + (255 - g) * fator); b = int(b + (255 - b) * fator)
        return f"{r:02X}{g:02X}{b:02X}"

    # Cor POR SERVIÇO: cada bloco tem uma cor-base; dentro do bloco, o 1º serviço fica na cor
    # cheia (mais escura) e os demais em tons progressivamente mais claros da MESMA cor.
    # Ex.: Enf+PA → Enfermaria = azul escuro · Enf+PA → PA mandic = azul claro.
    loc_cor_map = {}
    for i, l in enumerate(locais_cfg):
        base = cores_loc[i % len(cores_loc)]
        servs = [l] + (l.get("servicos_extras") or [])
        # rótulo do bloco (sem serviço específico) recebe a cor-base
        for r in [l.get("nome_bloco",""), l.get("abrev","")]:
            if r and r not in loc_cor_map:
                loc_cor_map[r] = base
        for j, se in enumerate(servs):
            shade = base if j == 0 else _clarear(base, min(0.30 + 0.22 * j, 0.78))
            for r in [se.get("nome",""), se.get("abrev","")]:
                if r and r not in loc_cor_map:
                    loc_cor_map[r] = shade
    # casa rótulos mais longos primeiro (ex: "Enfermaria"/"PA mandic" antes de "Enf+PA")
    loc_cor_itens = sorted(loc_cor_map.items(), key=lambda kv: -len(kv[0]))

    for ri, sem in enumerate(calendario):
        row = ri + 4
        sem_num = sem.get("semana", ri+1)
        periodo = sem.get("periodo","")
        if not periodo and ri < len(semanas):
            d0 = semanas[ri][0]; d1 = semanas[ri][-1]
            periodo = f"{d0.strftime('%d/%m')}–{d1.strftime('%d/%m/%Y')}"
        _cel(ws, row, 1, f"Sem {sem_num}", bold=True, bg="F2F2F2", sz=9)
        _cel(ws, row, 2, periodo, bg="F2F2F2", sz=9)
        aloc = sem.get("alocacao",{})
        for i, sg in enumerate(sgs):
            local = aloc.get(f"SG{sg}", aloc.get(sg, ""))
            cor = next((c for n,c in loc_cor_itens if n.lower() in local.lower()), "F2F2F2") if local else "F2F2F2"
            _cel(ws, row, 3+i, local, bg=cor, sz=9)

    # Blocos
    row = len(calendario) + 5
    _header(ws, row, 1, "BLOCOS DE RODÍZIO", span=2+n_sgs, bg=C["H2"])
    for bl in config.get("blocos",[]):
        row += 1
        _cel(ws, row, 1, bl.get("nome",""), bold=True, bg="F2F2F2", sz=9)
        _cel(ws, row, 2, bl.get("periodo",""), bg="F2F2F2", sz=9)
        _cel(ws, row, 3, bl.get("descricao",""), halign="left", bg="F2F2F2", sz=9, wrap=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=2+n_sgs)


# ── ABA 4: ESCALA NOMINAL DETALHADA ──────────────────────────────────────────
def _aba_escala_nominal(wb, titulo, escala_det, config, semanas, locais_cfg, ano_escala="2026"):
    ws = wb.create_sheet("Escala Nominal Detalhada")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E4"

    grupo = config.get("grupo","")
    turma = config.get("turma","")
    d_ini = semanas[0][0].strftime("%d/%m/%Y") if semanas else ""
    d_fim = semanas[-1][-1].strftime("%d/%m/%Y") if semanas else ""
    regras = config.get("regras_especiais",{})

    subtitulo = f"Terça = {regras.get('terca','12-16h')}  ·  Quinta = {regras.get('quinta','ENAMED')}"

    _header(ws, 1, 1, f"ESCALA NOMINAL DETALHADA — {grupo} / {turma}  |  {d_ini}–{d_fim}  |  rodízio equilibrado ✓",
            span=10, sz=11)
    _header(ws, 2, 1, subtitulo, span=10, bg=C["H2"], sz=9)

    for i, h in enumerate(["Sem","Data","Dia","Local","Turno","Horário","h","SG","Nome","RA"],1):
        _cel(ws, 3, i, h, bold=True, bg=C["H2"], fc="FFFFFF", sz=9)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["H"].width = 5
    ws.column_dimensions["I"].width = 36
    ws.column_dimensions["J"].width = 14

    sem_atual = None
    for entry in escala_det:
        sem = entry.get("semana","")
        if sem != sem_atual:
            sem_atual = sem
            # Linha separadora de semana
            periodo = entry.get("data","")
            label = f"SEMANA {sem}"
            if semanas and isinstance(sem, int) and 1 <= sem <= len(semanas):
                d0 = semanas[sem-1][0]; d1 = semanas[sem-1][-1]
                label = f"SEMANA {sem}  —  {d0.strftime('%d/%m')} a {d1.strftime('%d/%m/%Y')}"
            row = ws.max_row + 1
            _header(ws, row, 1, label, span=10, bg=C["H2"], sz=9)

        row = ws.max_row + 1
        local = entry.get("local","")
        turno = entry.get("turno","")
        data_raw = str(entry.get("data",""))
        data_norm = _normalizar_data(data_raw, ano_escala if 'ano_escala' in dir() else "2026")
        cor = _cor_local(local, locais_cfg)

        # Cor especial para FDS e cinderela
        turno_lower = turno.lower() if turno else ""
        eh_plantao = bool(entry.get("plantao"))
        if eh_plantao:
            cor = C["FDS_PLT"]
            turno = f"★ {turno}"   # plantão de complemento (em outro serviço)
        elif "fds" in turno_lower or "★" in turno or "plantão" in turno_lower:
            cor = C["FDS_PLT"]
        elif "reduz" in turno_lower or turno.upper() == "R":
            cor = C["TARDE_R"]
        elif "verde" in turno_lower or turno.upper() == "AV":
            cor = C["VERDE"]

        alunos = entry.get("alunos",[])
        nome_str = entry.get("nome","") if isinstance(entry.get("nome"), str) else ""
        if alunos and not nome_str:
            nome_str = " | ".join(alunos)

        vals = [sem, data_norm, entry.get("dia",""), local,
                turno, entry.get("horario",""), entry.get("horas",""),
                entry.get("sg",""), nome_str, entry.get("ra","")]
        for ci, v in enumerate(vals, 1):
            _cel(ws, row, ci, v, bg=cor, sz=8,
                 halign="left" if ci in [4,9] else "center", wrap=(ci==9))
        ws.row_dimensions[row].height = 13


# ── ABA 5: RESUMO DE HORAS ───────────────────────────────────────────────────
def _aba_resumo_horas(wb, titulo, resumo_horas, config, semanas, locais_cfg, dados=None):
    ws = wb.create_sheet("Resumo de Horas")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D3"

    grupo = config.get("grupo","")
    turma = config.get("turma","")
    n_sem = len(semanas)

    _header(ws, 1, 1, f"RESUMO DE HORAS — {grupo} / {turma}", span=5+n_sem, sz=11)

    # Headers semanas com data
    sem_labels = []
    for i, sem in enumerate(semanas):
        d = sem[0].strftime("%d/%m")
        sem_labels.append(f"S{i+1}\n{d}")

    headers = ["SG","Nome","RA"] + sem_labels + ["TOTAL","Cind","Períodos\nEnf FDS","Períodos\nPS FDS"]
    for i, h in enumerate(headers, 1):
        bg = C["H1"] if h in ["SG","TOTAL"] else (C["ENF_FDS_H"] if "Enf" in h else (C["PS_FDS_H"] if "PS" in h else C["H2"]))
        fc = "FFFFFF" if bg in [C["H1"],C["H2"]] else "000000"
        _cel(ws, 2, i, h, bold=True, bg=bg, fc=fc, sz=9, wrap=True)
        ws.row_dimensions[2].height = 28

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    for i in range(n_sem):
        ws.column_dimensions[get_column_letter(4+i)].width = 8
    ws.column_dimensions[get_column_letter(4+n_sem)].width = 10
    for j in range(3):
        ws.column_dimensions[get_column_letter(5+n_sem+j)].width = 12

    alunos_por_sg = config.get("alunos_por_sg",{})
    escala_det = dados.get("escala_detalhada", []) if dados else []
    ano_escala = "2026"
    try: ano_escala = date.fromisoformat(config.get("data_inicio","")).strftime("%Y")
    except: pass

    # Calcular horas reais por aluno por semana a partir da escala_detalhada
    # Com cap diário: máximo de horas por aluno por dia = limite_ch / 5
    from collections import defaultdict
    limite_ch_cfg = int(config.get("regras_especiais",{}).get("limite_ch", 40))
    limite_abs_cfg = int(config.get("regras_especiais",{}).get("limite_abs", 43))

    horas_calc = defaultdict(lambda: defaultdict(float))   # {nome: {sem_num: horas}}
    horas_dia  = defaultdict(lambda: defaultdict(float))   # {nome: {(sem,data): horas}} — para cap diário
    cind_calc = defaultdict(int)
    fds_enf_calc = defaultdict(int)
    fds_ps_calc = defaultdict(int)

    for entry in escala_det:
        alunos = entry.get("alunos", [])
        if isinstance(entry.get("nome"), str) and entry.get("nome"):
            alunos = [entry["nome"]]
        sem_num = entry.get("semana", 0)
        data_key = (int(sem_num), entry.get("data",""))
        horas_entry = entry.get("horas", 0)
        try: horas_entry = float(horas_entry)
        except: horas_entry = 0
        turno = (entry.get("turno","") or "").lower()
        local = (entry.get("local","") or "").lower()
        dia = (entry.get("dia","") or "").lower()
        # fim de semana: pelo dia (sáb/dom) ou pela data; o gerador marca FDS pelo DIA, não no turno
        eh_fds = dia.startswith("sáb") or dia.startswith("sab") or dia.startswith("dom") or "fds" in turno or "★" in turno
        if not eh_fds:
            try:
                dd, mm = str(entry.get("data","")).split("/")[:2]
                dt_fds = date(int(ano_escala), int(mm), int(dd))
                eh_fds = dt_fds.weekday() >= 5
            except Exception:
                pass
        for nome in alunos:
            horas_calc[nome][int(sem_num)] += horas_entry
            horas_dia[nome][data_key] += horas_entry
            if "cinderela" in turno or turno == "c":
                cind_calc[nome] += 1
            if eh_fds:
                if "enf" in local: fds_enf_calc[nome] += 1
                else: fds_ps_calc[nome] += 1

    # Montar mapa de resumo_horas por nome (para RA e dados extras)
    rh_map = {r.get("nome",""): r for r in resumo_horas} if resumo_horas else {}

    row = 3
    for sg_key in sorted(alunos_por_sg.keys(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        sg_num = int(sg_key) if sg_key.isdigit() else 1
        cor_sg = _cor_sg(sg_num)
        for i_al, nome in enumerate(alunos_por_sg[sg_key]):
            cor_row = C["ALT1"] if i_al % 2 == 0 else C["ALT2"]
            rh = rh_map.get(nome, {})

            # Usar horas calculadas da escala_detalhada; fallback para dados da IA
            semanas_h_calc = [horas_calc[nome].get(i+1, 0) for i in range(n_sem)]
            semanas_h_ia = rh.get("semanas", [0]*n_sem)

            # Preferir calculado se tiver dados, senão usar IA
            semanas_h = []
            for i in range(n_sem):
                h_calc = semanas_h_calc[i]
                h_ia = semanas_h_ia[i] if i < len(semanas_h_ia) else 0
                semanas_h.append(h_calc if h_calc > 0 else h_ia)

            total = sum(semanas_h)
            if total == 0:
                total = rh.get("total_horas", 0)

            _cel(ws, row, 1, f"SG{sg_num}", bold=True, bg=cor_sg, sz=9)
            _cel(ws, row, 2, nome, halign="left", bg=cor_row, sz=9)
            _cel(ws, row, 3, str(rh.get("ra","")), bg=cor_row, sz=9)

            limite_ch = int(config.get("regras_especiais",{}).get("limite_ch",40))
            limite_abs = int(config.get("regras_especiais",{}).get("limite_abs",43))

            for i_sem in range(n_sem):
                h = semanas_h[i_sem]
                cor_h = "FFC7CE" if h > limite_abs else ("FFEB9C" if h > limite_ch else cor_row)
                _cel(ws, row, 4+i_sem, f"{int(h)}h" if h else "—", bg=cor_h, sz=9)

            _cel(ws, row, 4+n_sem, f"{int(total)}h", bold=True, bg=C["TOTAL"], sz=9)
            cind_v = cind_calc.get(nome, rh.get("cinderelas",0))
            enf_v = fds_enf_calc.get(nome, rh.get("plantoes_enf_fds",0))
            ps_v = fds_ps_calc.get(nome, rh.get("plantoes_ps_fds",0))
            _cel(ws, row, 5+n_sem, str(cind_v), bg=cor_row, sz=9)
            _cel(ws, row, 6+n_sem, str(enf_v), bg=C["ENF_FDS_H"], sz=9)
            _cel(ws, row, 7+n_sem, str(ps_v), bg=C["PS_FDS_H"], sz=9)
            ws.row_dimensions[row].height = 15
            row += 1

    # Rodapé
    row += 1
    nota = f"CH padrão: {config.get('regras_especiais',{}).get('limite_ch',40)}h/sem | Absoluto: {config.get('regras_especiais',{}).get('limite_abs',43)}h | Vermelho > 43h | Amarelo > 40h"
    _cel(ws, row, 1, nota, halign="left", italic=True, bg="F2F2F2", sz=8, border=False)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4+n_sem)

    # Aviso de semanas com excesso de horas
    excedentes = []
    for nome_exc, sems in horas_calc.items():
        for sem_exc, h_exc in sems.items():
            if h_exc > limite_abs_cfg:
                excedentes.append(f"{nome_exc} (S{sem_exc}: {int(h_exc)}h)")
    if excedentes:
        row += 1
        aviso = f"⚠️ ATENÇÃO — Horas acima do limite ({limite_abs_cfg}h): " + " | ".join(excedentes[:10])
        _cel(ws, row, 1, aviso, halign="left", italic=True, bg="FFC7CE", sz=8, bold=True, border=False)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4+n_sem)

def _cor_semana_aluno(nome, sem_num, resumo_horas, locais_cfg):
    """Tenta retornar cor do local do aluno naquela semana."""
    return None  # simplificado — pode ser expandido


# ── ABA 6: REGRAS E RESTRIÇÕES ────────────────────────────────────────────────
def _aba_regras(wb, titulo, config, dados):
    ws = wb.create_sheet("Regras e Restrições")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 55

    esp = config.get("especialidade","")
    grupo = config.get("grupo","")
    turma = config.get("turma","")
    locais_cfg = config.get("locais",[])
    regras = config.get("regras_especiais",{})

    _header(ws, 1, 1, f"REGRAS E RESTRIÇÕES — {esp} — {grupo} / {turma}", span=3, sz=11)
    for i, h in enumerate(["CATEGORIA","ITEM","DETALHE"],1):
        _cel(ws, 2, i, h, bold=True, bg=C["H2"], fc="FFFFFF", sz=9)

    linhas = [
        ("RODÍZIO", "Descrição", config.get("rodizio_desc","")),
        ("QUINTA-FEIRA", "Regra", regras.get("quinta","Sem tarde — ENAMED")),
        ("TERÇA-FEIRA", "Regra", regras.get("terca","Tarde encurtada")),
        ("CH", "Limite padrão", f"{regras.get('limite_ch',40)}h/semana"),
        ("CH", "Limite absoluto", f"{regras.get('limite_abs',43)}h/semana"),
        ("FDS", "Regras", regras.get("fds","")),
    ]
    for l in locais_cfg:
        bloqs = l.get("bloqueios_tarde",[])
        bloq_str = " | ".join([f"{b['dia']}: {b['tipo']}{' '+b.get('horario','') if b.get('horario') else ''}" for b in bloqs])
        linhas.append((l.get("nome",""), "Manhã", f"{l.get('manha','')} | Mín:{l.get('min_manha',0)} Máx:{l.get('max_manha',0)}"))
        linhas.append((l.get("nome",""), "Tarde", f"{l.get('tarde','')} | Mín:{l.get('min_tarde',0)} Máx:{l.get('max_tarde',0)}"))
        if bloqs: linhas.append((l.get("nome",""), "Bloqueios tarde", bloq_str))
        if l.get("cinderela"): linhas.append((l.get("nome",""), "Cinderela", l.get("cinderela","")))
        if l.get("fds"):
            linhas.append((l.get("nome",""), "FDS — quem faz", l.get("fds_quem","")))
            linhas.append((l.get("nome",""), "FDS — compensação", l.get("fds_comp","")))

    cores_cat = {}; cores_cycle = [C["ENF"],C["PS"],C["AMB"],C["LOC4"],C["LOC5"],"F2F2F2"]
    cor_i = 0
    row = 3
    for cat, item, det in linhas:
        if not det: continue
        if cat not in cores_cat:
            cores_cat[cat] = cores_cycle[cor_i % len(cores_cycle)]; cor_i+=1
        cor = cores_cat[cat]
        _cel(ws, row, 1, cat, bold=True, bg=cor, halign="left", sz=9)
        _cel(ws, row, 2, item, bg=cor, halign="left", sz=9)
        _cel(ws, row, 3, det, bg=cor, halign="left", sz=9, wrap=True)
        ws.row_dimensions[row].height = 15
        row += 1


# ── ABA 7: ESCALA POR SUBGRUPO ────────────────────────────────────────────────
def _aba_escala_subgrupo(wb, titulo, alunos_por_sg, escala_por_aluno, config, semanas, locais_cfg):
    ws = wb.create_sheet("Escala por Subgrupo")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C6"

    grupo = config.get("grupo","")
    turma = config.get("turma","")
    d_ini = semanas[0][0].strftime("%d/%m/%Y") if semanas else ""
    d_fim = semanas[-1][-1].strftime("%d/%m/%Y") if semanas else ""

    _header(ws, 1, 1, f"ESCALA POR SUBGRUPO — {grupo} / {turma}  |  {d_ini}–{d_fim}  |  rodízio equilibrado",
            span=60, sz=11)
    _header(ws, 2, 1, "M=Manhã · T=Tarde · R=Tarde12-16h · F=FDS manhã · ★=Plantão de complemento (outro serviço)",
            span=60, bg=C["H2"], sz=9)

    # Montar todas as datas
    todas_datas = []
    for sem in semanas:
        todas_datas.extend(sem)
    n_datas = len(todas_datas)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    for i in range(n_datas):
        ws.column_dimensions[get_column_letter(3+i)].width = 9

    # Linha 3: Semanas (agrupadas)
    _cel(ws, 3, 1, "SG", bold=True, bg=C["H1"], fc="FFFFFF", sz=9)
    _cel(ws, 3, 2, "Nome", bold=True, bg=C["H1"], fc="FFFFFF", sz=9)
    col = 3
    for i_sem, sem in enumerate(semanas):
        d0 = sem[0]; d1 = sem[-1]
        label = f"S{i_sem+1}|{d0.strftime('%d/%m')}–{d1.strftime('%d/%m')}"
        _header(ws, 3, col, label, span=7, bg=C["H1"], sz=8)
        col += 7

    # Linha 4: Datas
    _cel(ws, 4, 1, "", bg=C["H2"])
    _cel(ws, 4, 2, "", bg=C["H2"])
    for i, dt in enumerate(todas_datas):
        eh_fds = dt.weekday() >= 5
        bg = C["FDS"] if eh_fds else C["H3"]
        _cel(ws, 4, 3+i, dt.strftime("%d/%m"), bold=True, bg=bg, sz=8)

    # Linha 5: Dias da semana
    _cel(ws, 5, 1, "", bg=C["H2"])
    _cel(ws, 5, 2, "", bg=C["H2"])
    for i, dt in enumerate(todas_datas):
        eh_fds = dt.weekday() >= 5
        bg = C["FDS"] if eh_fds else C["H3"]
        _cel(ws, 5, 3+i, DIAS_PT[dt.weekday()], bold=True, bg=bg, sz=8)

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 14

    row = 6
    for sg_key in sorted(alunos_por_sg.keys(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        sg_num = int(sg_key) if sg_key.isdigit() else 1
        cor_sg = _cor_sg(sg_num)
        nomes = alunos_por_sg[sg_key]

        # Linha de separador do SG
        label_sg = f"SG{sg_num} — " + " · ".join([n.split()[0] for n in nomes[:4]])
        _header(ws, row, 1, label_sg, span=2+n_datas, bg=C["H2"], sz=9)
        row += 1

        for nome in nomes:
            _cel(ws, row, 1, f"SG{sg_num}", bold=True, bg=cor_sg, sz=8)
            _cel(ws, row, 2, nome, halign="left", bg=cor_sg, sz=8)

            for i, dt in enumerate(todas_datas):
                data_str = dt.strftime("%d/%m/%Y")
                data_str2 = dt.strftime("%d/%m")
                data_str3 = dt.strftime("%Y-%m-%d")
                aluno_datas = escala_por_aluno.get(nome, {})
                valor = (aluno_datas.get(data_str) or
                         aluno_datas.get(data_str2) or
                         aluno_datas.get(data_str3) or "")
                eh_fds = dt.weekday() >= 5

                # Cor da célula
                if not valor:
                    if eh_fds:
                        valor = "—"; cor_cel = C["FDS_CELL"]
                    else:
                        valor = "Área Verde"; cor_cel = C["VERDE"]  # dia útil sem turno
                else:
                    local_nome = valor.split("(")[0] if "(" in valor else valor
                    cor_cel = _cor_local(local_nome, locais_cfg)
                    if "FDS" in valor or "★" in valor: cor_cel = C["FDS_PLT"]
                    elif "(R)" in valor or "+R" in valor: cor_cel = C["TARDE_R"]

                _cel(ws, row, 3+i, valor, bg=cor_cel, sz=7)

            ws.row_dimensions[row].height = 13
            row += 1


# ── ABA 8: ESCALA INDIVIDUAL (sub-colunas M/T/C por dia) ──────────────────────
def _aba_escala_individual(wb, titulo, alunos_por_sg, escala_por_aluno_per, config, semanas, locais_cfg):
    ws = wb.create_sheet("Escala Individual")
    ws.sheet_view.showGridLines = False

    grupo = config.get("grupo","")
    turma = config.get("turma","")
    d_ini = semanas[0][0].strftime("%d/%m/%Y") if semanas else ""
    d_fim = semanas[-1][-1].strftime("%d/%m/%Y") if semanas else ""

    _header(ws, 1, 1, f"ESCALA INDIVIDUAL — {grupo} / {turma}  |  {d_ini}–{d_fim}  |  por período (M/T/C)",
            span=80, sz=11)
    _header(ws, 2, 1, "M=Manhã · T=Tarde · C=Cinderela/Noturno · ★=Plantão  |  'Área Verde' = período de DIA ÚTIL livre "
                      "(obrigatório ≥1 por semana)", span=80, bg=C["H2"], sz=9)

    todas_datas = []
    for sem in semanas:
        todas_datas.extend(sem)
    n_datas = len(todas_datas)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 12
    for i in range(n_datas * 3):
        ws.column_dimensions[get_column_letter(4 + i)].width = 8

    # Linha 3: Nome/SG/RA (mesclados 3-4) + data (mesclada nas 3 sub-colunas).  Linha 4: M/T/C
    for ci, h in enumerate(["Nome", "SG", "RA"], 1):
        _cel(ws, 3, ci, h, bold=True, bg=C["H1"], fc="FFFFFF", sz=9)
        _cel(ws, 4, ci, "", bg=C["H1"])
        ws.merge_cells(start_row=3, start_column=ci, end_row=4, end_column=ci)
    for i, dt in enumerate(todas_datas):
        eh_fds = dt.weekday() >= 5
        bg = C["FDS"] if eh_fds else C["H3"]
        c0 = 4 + i * 3
        _cel(ws, 3, c0, dt.strftime("%d/%m"), bold=True, bg=bg, sz=8)
        for _cc in (c0 + 1, c0 + 2):
            _cel(ws, 3, _cc, "", bg=bg)
        ws.merge_cells(start_row=3, start_column=c0, end_row=3, end_column=c0 + 2)
        for j, lbl in enumerate(["M", "T", "C"]):
            _cel(ws, 4, c0 + j, lbl, bold=True, bg=bg, sz=7)
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 13
    ws.freeze_panes = "D5"

    ra_map = config.get("ra_por_aluno", {}) or {}
    row = 5
    for sg_key in sorted(alunos_por_sg.keys(), key=lambda x: int(x[0]) if x[0].isdigit() else 0):
        sg_num = int(sg_key) if sg_key.isdigit() else 1
        cor_sg = _cor_sg(sg_num)

        for i_al, nome in enumerate(alunos_por_sg[sg_key]):
            cor_nome = C["ALT1"] if i_al % 2 == 0 else C["ALT2"]
            _cel(ws, row, 1, nome, halign="left", bg=cor_sg, bold=True, sz=9)
            _cel(ws, row, 2, f"SG{sg_num}", bold=True, bg=cor_sg, sz=9)
            _cel(ws, row, 3, str(ra_map.get(str(nome).strip(), "")), bg=cor_nome, sz=9)

            dd = escala_por_aluno_per.get(nome, {})
            for i, dt in enumerate(todas_datas):
                per = (dd.get(dt.strftime("%d/%m/%Y")) or dd.get(dt.strftime("%d/%m"))
                       or dd.get(dt.strftime("%Y-%m-%d")) or {})
                eh_fds = dt.weekday() >= 5
                c0 = 4 + i * 3
                for j, b in enumerate(["M", "T", "C"]):
                    info = per.get(b)
                    if info:
                        txt = ("★" if info["plantao"] else "") + info["abrev"]
                        cor_cel = _cor_local(info["local"], locais_cfg)
                        if info["plantao"]:
                            cor_cel = C["FDS_PLT"]
                        elif info["reduz"]:
                            cor_cel = C["TARDE_R"]
                    elif b == "C":                       # noite/cinderela vazio NÃO é área verde
                        txt = ""; cor_cel = C["FDS_CELL"] if eh_fds else C["CIND"]
                    elif eh_fds:
                        txt = "—"; cor_cel = C["FDS_CELL"]
                    else:
                        txt = "Área Verde"; cor_cel = C["VERDE"]   # manhã/tarde de dia útil livre
                    _cel(ws, row, c0 + j, txt, bg=cor_cel, sz=7)

            ws.row_dimensions[row].height = 13
            row += 1

# v2 - formato CM5 exato


# ── ABA: ESCALA POR SERVIÇO ───────────────────────────────────────────────────
def _aba_por_servico(wb, titulo, escala_det, config, semanas, locais_cfg, ano_escala="2026"):
    ws = wb.create_sheet("Escala por Serviço")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C5"

    grupo = config.get("grupo","")
    turma = config.get("turma","")
    d_ini = semanas[0][0].strftime("%d/%m/%Y") if semanas else ""
    d_fim = semanas[-1][-1].strftime("%d/%m/%Y") if semanas else ""
    regras = config.get("regras_especiais",{})

    # Montar todas as datas
    todas_datas = [dt for sem in semanas for dt in sem]
    n_datas = len(todas_datas)

    _header(ws, 1, 1, f"ESCALA POR SERVIÇO — {grupo} / {turma}  |  {d_ini}–{d_fim}", span=2+n_datas, sz=11)
    subtit = f"Terça = {regras.get('terca','12-16h')}  ·  Quinta = {regras.get('quinta','ENAMED')}"
    _header(ws, 2, 1, subtit, span=2+n_datas, bg=C["H2"], sz=9)

    # Linha 3: datas
    _cel(ws, 3, 1, "Local", bold=True, bg=C["H1"], fc="FFFFFF", sz=9)
    _cel(ws, 3, 2, "#", bold=True, bg=C["H1"], fc="FFFFFF", sz=9)
    for i, dt in enumerate(todas_datas):
        eh_fds = dt.weekday() >= 5
        eh_ter = dt.weekday() == 1
        bg = C["FDS"] if eh_fds else (C["TER"] if eh_ter else C["H3"])
        lbl = dt.strftime("%d/%m")
        _cel(ws, 3, 3+i, lbl, bold=True, bg=bg, sz=8)

    # Linha 4: dias com anotações especiais
    _cel(ws, 4, 1, "", bg=C["H1"])
    _cel(ws, 4, 2, "", bg=C["H1"])
    for i, dt in enumerate(todas_datas):
        eh_fds = dt.weekday() >= 5
        eh_ter = dt.weekday() == 1
        eh_qui = dt.weekday() == 3
        bg = C["FDS"] if eh_fds else (C["TER"] if eh_ter else C["H3"])
        dia = DIAS_PT[dt.weekday()]
        if eh_ter: dia += " ★16h"
        if eh_qui: dia = "Qui/ENAMED"
        _cel(ws, 4, 3+i, dia, bold=True, bg=bg, sz=8)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 6
    for i in range(n_datas):
        ws.column_dimensions[get_column_letter(3+i)].width = 11
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 14

    # Montar índice de quem está em cada (data, local, turno) → lista de alunos
    from collections import defaultdict
    # {(data_norm, local, turno): [alunos]}
    grade = defaultdict(list)
    for entry in escala_det:
        data_raw = str(entry.get("data",""))
        data_n = _normalizar_data(data_raw, ano_escala)
        local = entry.get("local","")
        turno = entry.get("turno","")
        alunos = entry.get("alunos", [])
        if isinstance(entry.get("nome"), str) and entry.get("nome"):
            alunos = [entry["nome"]]
        hor = entry.get("horario","")
        plt = bool(entry.get("plantao"))
        for a in alunos:
            grade[(data_n, local, turno)].append((a, hor, plt))

    # Agrupar turnos por local
    locais_nomes = list(dict.fromkeys([entry.get("local","") for entry in escala_det if entry.get("local")]))

    row = 5
    for li, local in enumerate(locais_nomes):
        cor_loc = _cor_local(local, locais_cfg)

        # Descobrir turnos deste local, em ordem natural: Manhã → Tarde → Cinderela
        turnos_local = list(dict.fromkeys([
            entry.get("turno","") for entry in escala_det
            if entry.get("local","") == local and entry.get("turno")
        ]))
        _ordem_turno = {"manhã": 0, "manha": 0, "tarde": 1, "cinderela": 2, "cind": 2, "noite": 3}
        turnos_local.sort(key=lambda t: _ordem_turno.get(str(t).strip().lower(), 9))

        # Separador fino entre LOCAIS (não entre turnos) — evita repetir o nome
        if li > 0:
            for col in range(1, 3+n_datas):
                _cel(ws, row, col, "", bg="D0D9E8", border=False)
            ws.row_dimensions[row].height = 6
            row += 1

        bloco_inicio = row  # 1ª linha do conteúdo deste local (para a faixa lateral)

        for turno in turnos_local:
            # Header do turno (o nome do local NÃO se repete aqui — vira faixa lateral)
            for i, dt in enumerate(todas_datas):
                eh_fds = dt.weekday() >= 5
                eh_ter = dt.weekday() == 1
                eh_qui = dt.weekday() == 3
                bg_h = C["FDS"] if eh_fds else (C["TER"] if eh_ter else C["H2"])
                # Verificar se há ENAMED (quinta)
                if eh_qui and regras.get("quinta",""):
                    lbl_h = "ENAMED"
                    bg_h = C["ENAMED"]
                else:
                    lbl_h = {"cinderela": "CIND", "cind": "CIND"}.get(str(turno).strip().lower(), turno.upper()[:5])
                # texto preto em fundos claros (FDS/terça/ENAMED); branco só no azul do header
                fc_h = "FFFFFF" if bg_h == C["H2"] else "000000"
                _cel(ws, row, 3+i, lbl_h, bold=True, bg=bg_h, fc=fc_h, sz=8)
            _cel(ws, row, 2, turno[:10], bold=True, bg=C["H2"], fc="FFFFFF", sz=8)
            row += 1

            # Descobrir máx de alunos por turno/dia para este local
            max_alunos = 0
            for dt in todas_datas:
                data_n = dt.strftime("%d/%m/%Y")
                lst = grade.get((data_n, local, turno), [])
                if not lst:
                    data_n2 = dt.strftime("%d/%m")
                    lst = grade.get((data_n2, local, turno), [])
                max_alunos = max(max_alunos, len(lst))
            max_alunos = max(max_alunos, 1)

            # Linhas de alunos (slots)
            for slot in range(max_alunos):
                # Linha do nome (coluna A fica vazia — pertence à faixa do local)
                _cel(ws, row, 2, f"#{slot+1}", bold=True, bg="F2F2F2", sz=8)
                for i, dt in enumerate(todas_datas):
                    eh_fds = dt.weekday() >= 5
                    eh_ter = dt.weekday() == 1
                    eh_qui = dt.weekday() == 3
                    data_n = dt.strftime("%d/%m/%Y")
                    lst = grade.get((data_n, local, turno), [])
                    if not lst:
                        data_n2 = dt.strftime("%d/%m")
                        lst = grade.get((data_n2, local, turno), [])
                    eh_plt = slot < len(lst) and len(lst[slot]) > 2 and lst[slot][2]
                    if eh_fds:
                        bg_c = C["FDS"]
                        nome_c = lst[slot][0].split()[0] if slot < len(lst) else ""
                    elif eh_qui and not lst:
                        bg_c = C["ENAMED"]; nome_c = "ENAMED"
                    else:
                        bg_c = C["TER"] if eh_ter else "FFFFFF"
                        if slot < len(lst):
                            nome_c = lst[slot][0].split()[0] + " " + (lst[slot][0].split()[1][:1] + "." if len(lst[slot][0].split()) > 1 else "")
                        else:
                            bg_c = C["FDS"] if eh_fds else "F5F5F5"
                            nome_c = ""
                    if eh_plt and nome_c:
                        nome_c = "★ " + nome_c            # plantão de complemento
                        bg_c = C["FDS_PLT"]               # fundo lilás p/ destacar
                    _cel(ws, row, 3+i, nome_c, bg=bg_c, sz=8)
                ws.row_dimensions[row].height = 13
                row += 1

                # Linha do horário (coluna A fica vazia — pertence à faixa do local)
                _cel(ws, row, 2, "", bg="F2F2F2")
                for i, dt in enumerate(todas_datas):
                    eh_fds = dt.weekday() >= 5
                    eh_qui = dt.weekday() == 3
                    data_n = dt.strftime("%d/%m/%Y")
                    lst = grade.get((data_n, local, turno), [])
                    if not lst:
                        lst = grade.get((dt.strftime("%d/%m"), local, turno), [])
                    if eh_fds:
                        bg_h = C["FDS"]; hor_c = lst[slot][1] if slot < len(lst) else ""
                    elif eh_qui and not lst:
                        bg_h = C["ENAMED"]; hor_c = ""
                    else:
                        bg_h = C["H2"]
                        hor_c = lst[slot][1] if slot < len(lst) else ""
                    _cel(ws, row, 3+i, hor_c, bg=bg_h, fc="FFFFFF" if bg_h == C["H2"] else "000000", sz=7)
                ws.row_dimensions[row].height = 11
                row += 1

        # ── Faixa lateral do LOCAL: nome escrito UMA vez, mesclado na vertical ──
        bloco_fim = row - 1
        if bloco_fim >= bloco_inicio:
            ws.merge_cells(start_row=bloco_inicio, start_column=1, end_row=bloco_fim, end_column=1)
            cel_loc = _cel(ws, bloco_inicio, 1, local, bold=True, bg=cor_loc, fc="1F4E79", sz=11)
            cel_loc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, textRotation=90)

    return ws
