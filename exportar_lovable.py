# -*- coding: utf-8 -*-
"""Exporta a escala em 2 formatos que o Lovable importa:
  1) TEMPLATE EM LOTE  -> gerar_template_lovable(dados, config)
     Abas: Semana Padrão · Distribuição · _Serviços · _Blocos · _Subgrupos · LEIA-ME
  2) CORREÇÃO/IMPORTAR -> gerar_correcao_lovable(dados, config)
     Abas: Instruções · Correções <rodízio>  (1 linha por aluno × dia com atividade)
"""
import io
import unicodedata
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DIAS3 = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
DIA_SEMANA_LONG = ["segunda", "terca", "quarta", "quinta", "sexta", "sabado", "domingo"]
AMARELO = PatternFill("solid", fgColor="FFF2CC")
CINZA = PatternFill("solid", fgColor="D9D9D9")
HDR = PatternFill("solid", fgColor="1F4E78")
HDRF = Font(bold=True, color="FFFFFF")


def _sa(s):
    return "".join(c for c in unicodedata.normalize("NFKD", str(s or "")) if not unicodedata.combining(c)).strip().lower()


def _normserv(s):
    """Normaliza nome de serviço p/ casamento: sem acento, minúsculo, pontuação -> espaço."""
    s = "".join(c for c in unicodedata.normalize("NFKD", str(s or "")) if not unicodedata.combining(c)).lower()
    return " ".join("".join(c if c.isalnum() else " " for c in s).split())


# ───────────── Catálogo de serviços cadastrados no Lovable (fonte: cadastro oficial) ─────────────
# (nome EXATO cadastrado no Lovable, especialidade)
_LOVABLE_SERVICOS = [
    ('AMB OFTALMO MANDIC', 'CLÍNICA CIRÚRGICA'),
    ('AMBULATÓRIO CIRURGIA HSLMA', 'CLÍNICA CIRÚRGICA'),
    ('ANESTESIO - SC LEME', 'CLÍNICA CIRÚRGICA'),
    ('ANESTESIO - SCA', 'CLÍNICA CIRÚRGICA'),
    ('ANESTESIO-MANDIC', 'CLÍNICA CIRÚRGICA'),
    ('CC MANDIC', 'CLÍNICA CIRÚRGICA'),
    ('CC SCA', 'CLÍNICA CIRÚRGICA'),
    ('ENF CIR SCA', 'CLÍNICA CIRÚRGICA'),
    ('ENFERMARIA CIRÚRGICA E CC-LEME', 'CLÍNICA CIRÚRGICA'),
    ('ORTOPEDIA SCA', 'CLÍNICA CIRÚRGICA'),
    ('PS SCA', 'CLÍNICA CIRÚRGICA'),
    ('SAMU', 'CLÍNICA CIRÚRGICA'),
    ('AMBULATÓRIO CLÍNICA MÉDICA-MANDIC', 'CLÍNICA MÉDICA'),
    ('ENF CM HSLMA', 'CLÍNICA MÉDICA'),
    ('ENFMANDIC', 'CLÍNICA MÉDICA'),
    ('PA ADULTO MANDIC', 'CLÍNICA MÉDICA'),
    ('Pronto Socorro SC LEME', 'CLÍNICA MÉDICA'),
    ('PS ADULTO SCA', 'CLÍNICA MÉDICA'),
    ('SALA EMERGÊNCIA- LEME', 'CLÍNICA MÉDICA'),
    ('UPA ADULTO MANDIC', 'CLÍNICA MÉDICA'),
    ('UTI HSLM', 'CLÍNICA MÉDICA'),
    ('AULA DE CORREÇÕES DE QUESTÕES - REFORÇO ENAMED', 'DIVERSAS'),
    ('AULA TEÓRICA DO MÓDULO', 'DIVERSAS'),
    ('AMBULATÓRIO GO', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('AULA TEÓRICA GO', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('CC GO MANDIC', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('CENTRO OBSTÉTRICO ARARAS', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('CENTRO OBSTÉTRICO LEME', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('CIRURGIA GINECOLÓGICA SCA', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('CO LEME', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('CO SCA', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('ENFERMARIA GO SCA', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('PNAR GO SCA', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('PS GO SCA', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('SIMULAÇÃO GO', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('UBS Alberto Franzini - Ginecologia e Obstetrícia', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('USG GO MANDIC', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('USG GO SCA', 'GINECOLOGIA E OBSTETRÍCIA'),
    ('ESF BENTO FERES', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('ESF EDMUNDO ULSON', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('ESF FERMIN BLANCO', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('ESF JERÔNYMO OMETTO', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('ESF LUCIA MENEGHETTI', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('ESF OPHELIA PESCE', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('ESF SIMÕES PONTES', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('LIVRE MFC', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('PSF NARCISO GOMES II', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('UBS', 'MEDICINA DE FAMÍLIA E COMUNIDADE'),
    ('AMBULATÓRIO PED HSLMA', 'PEDIATRIA'),
    ('AULA TEÓRICA PED 5º ANO (08-12H)', 'PEDIATRIA'),
    ('AULA TEÓRICA PED 6º ANO (10-12H)', 'PEDIATRIA'),
    ('CORREÇÃO DE QUESTÕES', 'PEDIATRIA'),
    ('ENF PED LEME', 'PEDIATRIA'),
    ('ENFERMARIA PED MANDIC', 'PEDIATRIA'),
    ('PA PED MANDIC', 'PEDIATRIA'),
    ('PNAR SCA', 'PEDIATRIA'),
    ('SALA PARTO E AC LEME', 'PEDIATRIA'),
    ('SC LIMEIRA PED', 'PEDIATRIA'),
    ('SIMULAÇÃO PED', 'PEDIATRIA'),
    ('UPA PED MANDIC', 'PEDIATRIA'),
    ('ESTÁGIO ELETIVO', 'SAÚDE MENTAL'),
    ('AMBULATÓRIO PSQ HSLMA', 'SAÚDE MENTAL (1 MÊS) + ELETIVO (1 MÊS)'),
    ('AULA TEÓRICA SM (08-10H)', 'SAÚDE MENTAL (1 MÊS) + ELETIVO (1 MÊS)'),
    ('ELETIVO', 'SAÚDE MENTAL (1 MÊS) + ELETIVO (1 MÊS)'),
    ('ENF QD + TRIAGEM PSQ HSLMA', 'SAÚDE MENTAL (1 MÊS) + ELETIVO (1 MÊS)'),
    ('ENFERMARIA PQS HSLMA', 'SAÚDE MENTAL (1 MÊS) + ELETIVO (1 MÊS)'),
    ('AMBULATÓRIO DEPENDÊNCIA QUÍMICA INFANTIL', 'SAÚDE MENTAL + ELETIVO'),
    ('ÁREA LIVRE', 'TODAS'),
    ('ÁREA VERDE', 'TODAS'),
    ('CIRURGIA SCA', 'TODAS'),
]

# Mapa: nome normalizado do serviço cadastrado -> nome EXATO (canônico) no Lovable.
_REG_BY_NORM = {_normserv(n): n for n, _e in _LOVABLE_SERVICOS}

# Apelidos: nomes/abreviações usados no app que NÃO batem 1:1 com o cadastro do Lovable.
_ALIAS_RAW = {
    # Clínica Cirúrgica
    "Enfermaria Cirúrgica e centro cirurgico SC LEME": "ENFERMARIA CIRÚRGICA E CC-LEME",
    "ENF CIR LEME": "ENFERMARIA CIRÚRGICA E CC-LEME",
    "Anestesiologia - SCA": "ANESTESIO - SCA",
    "ANESTESIO SCA": "ANESTESIO - SCA",
    "Anestesio Mandic": "ANESTESIO-MANDIC",
    "Anest mandic": "ANESTESIO-MANDIC",
    "Anestesio SC LEME": "ANESTESIO - SC LEME",
    "ANEST leme": "ANESTESIO - SC LEME",
    "ENFERMARIA/CC SCA": "ENF CIR SCA",
    "ENF/CC SCA": "ENF CIR SCA",
}
_ALIAS = {_normserv(k): v for k, v in _ALIAS_RAW.items()}


def _lovable_nome(nome):
    """Devolve o nome EXATO cadastrado no Lovable. Cai no original se não houver correspondência."""
    if not nome:
        return nome
    k = _normserv(nome)
    if k in _REG_BY_NORM:
        return _REG_BY_NORM[k]
    if k in _ALIAS:
        return _ALIAS[k]
    return nome


def _turno_key(t):
    t = _sa(t)
    if "manh" in t:
        return "manha"
    if "tard" in t:
        return "tarde"
    if "cind" in t or "noit" in t or "noturn" in t:
        return "cind"
    if "fds" in t:
        return "fds"
    return t


def _hhmm(p):
    """'07', '7', '07:00', '13h' -> 'HH:MM'."""
    p = str(p or "").strip().lower().replace("h", "").strip()
    if not p:
        return ""
    if ":" in p:
        a, b = p.split(":")[:2]
        return f"{int(a):02d}:{int(b):02d}"
    try:
        return f"{int(p):02d}:00"
    except ValueError:
        return p


def _horario_par(hor, tk):
    """'07-13h' / '07:00-12:00' -> ('07:00','13:00'). Cinderela -> overnight."""
    s = str(hor or "").strip()
    if "-" in s:
        a, b = s.split("-", 1)
        return _hhmm(a), _hhmm(b)
    # defaults por turno
    return {"manha": ("07:00", "12:00"), "tarde": ("13:00", "18:00"),
            "cind": ("19:00", "23:00")}.get(tk, ("", ""))


def _periodo_label(tk):
    return {"manha": "Manhã", "tarde": "Tarde", "cind": "Noturno"}.get(tk, "Manhã")


def _servicos_do_bloco(loc):
    return [loc] + (loc.get("servicos_extras") or [])


def _nome_bloco(loc, i):
    nb = loc.get("nome_bloco") or loc.get("nome") or f"Bloco {i+1}"
    # já vem no padrão "Bloco N-..."? se não, prefixa
    return nb if _sa(nb).startswith("bloco") else f"Bloco {i+1}-{nb}"


def _datas(data_inicio, num_semanas):
    try:
        d0 = date.fromisoformat(str(data_inicio)[:10])
    except Exception:
        d0 = date(2026, 7, 6)
    return [[d0 + timedelta(weeks=s, days=d) for d in range(7)] for s in range(int(num_semanas))]


# Calendário de rodízios do internato (T6 / 5º ano 2026.2). Comparação por (mês, dia).
_RODIZIOS = [
    (1, (7, 6),  (8, 30)),
    (2, (8, 31), (10, 25)),
    (3, (10, 26), (12, 20)),
    (4, (1, 4),  (2, 28)),
    (5, (3, 1),  (4, 25)),
    (6, (4, 26), (6, 20)),
]

# Códigos curtos de especialidade usados no código do rodízio (ex.: R1-PED-T6).
_ESP_COD = {
    "pediatria": "PED",
    "clinica medica": "CM", "clinica cirurgica": "CIR",
    "ginecologia e obstetricia": "GO", "ginecologia": "GO", "obstetricia": "GO",
    "saude mental": "SM", "saude mental + ferias": "SM",
    "medicina de familia e comunidade": "MFC", "mfc": "MFC",
}


def _num_rodizio(data_inicio):
    """Número do rodízio (1..6) a partir da data de início, conforme o calendário do internato."""
    try:
        d = date.fromisoformat(str(data_inicio)[:10])
    except Exception:
        return None
    md = (d.month, d.day)
    for n, ini, fim in _RODIZIOS:
        if ini <= md <= fim:
            return n
    return None


def _esp_cod(esp):
    e = _sa(esp)
    if e in _ESP_COD:
        return _ESP_COD[e]
    return next((v for k, v in _ESP_COD.items() if k in e or e in k), e.upper()[:3] or "ESC")


def _rod_codigo(config):
    esp = _esp_cod(config.get("especialidade", ""))
    turma = str(config.get("turma", "") or "").upper().replace(" ", "")
    n = _num_rodizio(config.get("data_inicio"))
    pref = f"R{n}-" if n else ""
    return f"{pref}{esp}-{turma}".strip("-") or "RODIZIO"


# ───────────────────────────── ARQUIVO 1: TEMPLATE ─────────────────────────────
def gerar_template_lovable(dados, config):
    wb = openpyxl.Workbook()
    locais = config.get("locais", [])
    num_sem = int(config.get("num_semanas", 8))
    alunos_por_sg = config.get("alunos_por_sg", {})
    ra_por_aluno = config.get("ra_por_aluno", {}) or {}
    det = dados.get("escala_detalhada") or []
    calend = dados.get("calendario_rodizio") or []

    nomes_bloco = [_nome_bloco(loc, i) for i, loc in enumerate(locais)]

    # ── Semana Padrão (1 linha por vaga: bloco × dia × turno × serviço) ──
    ws = wb.active
    ws.title = "Semana Padrão"
    ws.append(["Bloco", "Bloco Nome", "Dia", "Hora Início", "Hora Fim", "Período", "Serviço"])
    for c in range(1, 8):
        ws.cell(1, c).fill = HDR
        ws.cell(1, c).font = HDRF
    for i, loc in enumerate(locais):
        bnome = nomes_bloco[i]
        for s in _servicos_do_bloco(loc):
            snome = s.get("nome") or s.get("abrev") or ""
            if not snome:
                continue
            snome = _lovable_nome(snome)   # nome exato cadastrado no Lovable
            # turnos de dia útil
            for tk, hor, mx in (("manha", s.get("manha"), s.get("max_manha")),
                                ("tarde", s.get("tarde"), s.get("max_tarde")),
                                ("cind", s.get("cinderela"), s.get("max_cind"))):
                if not hor:
                    continue
                hi, hf = _horario_par(hor, tk)
                vagas = max(int(mx or 0), 1)
                dias = DIAS3[:5]
                if tk == "cind" and s.get("dias_cind"):
                    dias = [d for d in DIAS3[:5] if d in {x[:3].capitalize() for x in s.get("dias_cind")}] or DIAS3[:5]
                for dia in dias:
                    for _ in range(vagas):
                        ws.append([i + 1, bnome, dia, hi, hf, _periodo_label(tk), snome])
            # turnos de FDS
            for tk, hor, mx in (("manha", s.get("fds_manha"), s.get("fds_max_manha")),
                                ("tarde", s.get("fds_tarde"), s.get("fds_max_tarde")),
                                ("cind", s.get("fds_cind"), s.get("fds_max_cind"))):
                if not hor:
                    continue
                hi, hf = _horario_par(hor, tk)
                vagas = max(int(mx or 0), 1)
                for dia in ("Sáb", "Dom"):
                    for _ in range(vagas):
                        ws.append([i + 1, bnome, dia, hi, hf, _periodo_label(tk), snome])
    for col, w in zip("ABCDEFG", (7, 36, 6, 11, 10, 9, 34)):
        ws.column_dimensions[col].width = w

    # ── Distribuição (Subgrupo × Semana -> nome do bloco) ──
    wd = wb.create_sheet("Distribuição")
    wd.append(["Subgrupo"] + [f"Sem {s+1}" for s in range(num_sem)])
    for c in range(1, num_sem + 2):
        wd.cell(1, c).fill = HDR
        wd.cell(1, c).font = HDRF
    # casa o 'dest' do calendário (ex.: "ENF CIR LEME (Leme1)") com o bloco correto.
    # Espelha a resolução do escalonador: identifica o bloco pelo nome/abrev DELE e,
    # em seguida, por qualquer serviço (nome/abrev) que ele contenha.
    def _bloco_de_dest(dest):
        a = _normserv(dest)
        if not a:
            return ""
        for i, loc in enumerate(locais):
            ident = {_normserv(loc.get("nome_bloco")), _normserv(loc.get("nome")),
                     _normserv(loc.get("abrev"))} - {""}
            svc_match = any(
                x and x in a
                for s in _servicos_do_bloco(loc)
                for x in ({_normserv(s.get("nome")), _normserv(s.get("abrev"))} - {""})
            )
            if svc_match or any(x and x in a for x in ident):
                return nomes_bloco[i]
        return dest  # fallback: mantém o texto original

    # mapa (sg, semana) -> nome do bloco, a partir do calendário de rodízio
    blk_por_sg_sem = {}
    for entry in calend:
        sem = entry.get("semana")
        aloc = entry.get("alocacao", {}) or {}
        for sg_key, dest in aloc.items():
            sgn = "".join(ch for ch in str(sg_key) if ch.isdigit())
            if not sgn:
                continue
            blk_por_sg_sem[(sgn, int(sem))] = _bloco_de_dest(dest)
    for sg in sorted(alunos_por_sg.keys(), key=lambda x: int("".join(c for c in x if c.isdigit()) or 0)):
        sgn = "".join(c for c in sg if c.isdigit())
        wd.append([sgn] + [blk_por_sg_sem.get((sgn, s + 1), "") for s in range(num_sem)])
    wd.column_dimensions["A"].width = 10
    for s in range(num_sem):
        wd.column_dimensions[chr(66 + s)].width = 34

    # ── _Serviços ──  (Nome = nome EXATO cadastrado no Lovable)
    wsv = wb.create_sheet("_Serviços")
    wsv.append(["Nome", "Nome Curto", "nome do serviço cadastrado no lovable"])
    for c in range(1, 4):
        wsv.cell(1, c).fill = HDR
        wsv.cell(1, c).font = HDRF
    vistos = set()
    for loc in locais:
        for s in _servicos_do_bloco(loc):
            nm = s.get("nome") or s.get("abrev") or ""
            if not nm:
                continue
            lov = _lovable_nome(nm)
            if _sa(lov) not in vistos:
                vistos.add(_sa(lov))
                wsv.append([lov, s.get("abrev", ""), lov])
    for extra in ("ÁREA VERDE",):
        if _sa(extra) not in vistos:
            wsv.append([extra, "VERDE", "ÁREA VERDE"])
    wsv.column_dimensions["A"].width = 38
    wsv.column_dimensions["B"].width = 18
    wsv.column_dimensions["C"].width = 38

    # ── _Blocos ──
    wb_ = wb.create_sheet("_Blocos")
    wb_.append(["Índice", "Nome", "Semanas"])
    for i, loc in enumerate(locais):
        wb_.append([i + 1, nomes_bloco[i], int(loc.get("sem_por_sg", 0) or 0) or
                    max(1, num_sem // max(len(locais), 1))])
    wb_.column_dimensions["B"].width = 36

    # ── _Subgrupos ──
    wsg = wb.create_sheet("_Subgrupos")
    wsg.append(["Subgrupo", "Qtd Alunos", "Alunos (RA - Nome)"])
    for sg in sorted(alunos_por_sg.keys(), key=lambda x: int("".join(c for c in x if c.isdigit()) or 0)):
        nomes = alunos_por_sg[sg]
        linha = ["".join(c for c in sg if c.isdigit()), len(nomes)]
        for nome in nomes:
            ra = ra_por_aluno.get(nome, "")
            linha.append(f"{ra} - {nome}" if ra else nome)
        wsg.append(linha)
    wsg.column_dimensions["A"].width = 10
    wsg.column_dimensions["C"].width = 40

    # ── LEIA-ME ──
    wl = wb.create_sheet("LEIA-ME")
    cod = _rod_codigo(config)
    di = str(config.get("data_inicio", ""))[:10]
    for txt in [
        "INSTRUÇÕES — TEMPLATE EM LOTE",
        "",
        f"Rodízio: {cod} — {config.get('especialidade','')} ({config.get('turma','')})",
        f"Período: {di} — {num_sem} semana(s)",
        "",
        '1) Aba "Semana Padrão"',
        "   - Uma linha por slot (bloco × dia × período).",
        '   - Edite a coluna "Serviço" usando exatamente o Nome ou Nome Curto da aba "_Serviços".',
        '   - Pode adicionar/remover linhas livremente. O período (Manhã/Tarde/Noturno) é informativo; o bucket real é pela "Hora Início".',
        '   - Deixar "Serviço" em branco remove o slot.',
        '2) Aba "Distribuição"',
        "   - Matriz Subgrupo × Semana. Cada célula = NOME DO BLOCO (ver aba \"_Blocos\").",
        "   - Vazio = nenhum bloco para aquele subgrupo naquela semana.",
        "",
        "ATENÇÃO: Ao importar, TUDO da Semana Padrão e da Distribuição deste rodízio será SUBSTITUÍDO,",
        "e a escala diária (escalas_diarias) será regerada automaticamente.",
    ]:
        wl.append([txt])
    wl["A1"].font = Font(bold=True, size=12)
    wl.column_dimensions["A"].width = 110

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ───────────────────────── ARQUIVO 2: CORREÇÃO / IMPORTAR ─────────────────────
def gerar_correcao_lovable(dados, config):
    det = dados.get("escala_detalhada") or []
    alunos_por_sg = config.get("alunos_por_sg", {})
    ra_por_aluno = config.get("ra_por_aluno", {}) or {}
    num_sem = int(config.get("num_semanas", 8))
    semanas = _datas(config.get("data_inicio"), num_sem)

    sg_de = {}
    for sg, nomes in alunos_por_sg.items():
        for n in nomes:
            sg_de["".join(c for c in sg if c.isdigit())] = sg_de.get(sg, sg)
            sg_de[n] = "".join(c for c in sg if c.isdigit())

    # índice: (nome, 'DD/MM') -> {turno_key: serviço}
    idx = {}
    for e in det:
        data = str(e.get("data", ""))
        dd = data[:5] if len(data) >= 5 else data
        tk = _turno_key(e.get("turno"))
        serv = _lovable_nome(e.get("local", ""))   # nome exato cadastrado no Lovable
        for al in (e.get("alunos") or ([e["nome"]] if e.get("nome") else [])):
            idx.setdefault((al, dd), {})[tk] = serv

    wb = openpyxl.Workbook()
    cod = _rod_codigo(config)

    # ── Instruções ──
    wi = wb.active
    wi.title = "Instruções"
    di = semanas[0][0].strftime("%d/%m/%Y") if semanas else ""
    df = semanas[-1][-1].strftime("%d/%m/%Y") if semanas else ""
    servicos = sorted({_lovable_nome(s.get("nome") or s.get("abrev") or "")
                       for loc in config.get("locais", []) for s in _servicos_do_bloco(loc)} - {""})
    linhas = [
        f"Planilha de Correção — {cod} ({config.get('especialidade','')} {config.get('ano_curso','')} {config.get('turma','')})",
        "", f"Período: {di} a {df} ({num_sem} semanas)", "",
        "COMO USAR:",
        f'1. Vá para a aba "Correções {cod}".',
        "2. Cada linha = 1 aluno em 1 dia (manhã / tarde / noite).",
        '3. As colunas em AMARELO ("*_correto") trazem o serviço final — ajuste se algo estiver errado.',
        '4. Se a célula deve ficar SEM serviço, escreva "VAZIO".',
        "5. Use \"observacao\" para explicações pontuais.",
        "", "SERVIÇOS DISPONÍVEIS (use exatamente o nome):",
    ] + [f"• {s}" for s in servicos] + ["• ÁREA VERDE",
        "• Outro nome — avise no campo observacao para criar o serviço"]
    for t in linhas:
        wi.append([t])
    wi["A1"].font = Font(bold=True, size=12)
    wi.column_dimensions["A"].width = 90

    # ── Correções ──
    ws = wb.create_sheet(f"Correções {cod}"[:31])
    cols = ["data", "dia_semana", "subgrupo", "ra", "aluno",
            "manha_atual", "manha_correto", "tarde_atual", "tarde_correto",
            "noite_atual", "noite_correto", "observacao"]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(1, c).fill = HDR
        ws.cell(1, c).font = HDRF
    for c in (7, 9, 11):  # *_correto em amarelo
        ws.cell(1, c).fill = AMARELO
        ws.cell(1, c).font = Font(bold=True)

    todos = [(sg, n) for sg in sorted(alunos_por_sg, key=lambda x: int("".join(c for c in x if c.isdigit()) or 0))
             for n in alunos_por_sg[sg]]
    for semana in semanas:
        for dt in semana:
            dd = dt.strftime("%d/%m")
            eh_fds = dt.weekday() >= 5
            dia_long = DIA_SEMANA_LONG[dt.weekday()]
            for sg, nome in todos:
                serv_dia = idx.get((nome, dd), {})
                if not serv_dia:
                    continue   # só dias com atividade real
                sgn = "".join(c for c in sg if c.isdigit())
                ra = ra_por_aluno.get(nome, "")
                cel = {}
                for tk, col in (("manha", "manha"), ("tarde", "tarde"), ("cind", "noite")):
                    serv = serv_dia.get(tk, "")
                    atual = serv or None
                    if serv:
                        correto = serv
                    elif eh_fds or tk == "cind":
                        # FDS livre, ou período noturno sem plantão: vazio (NÃO é área verde —
                        # área verde é período de DIA ÚTIL livre, só manhã/tarde)
                        correto = None
                    else:
                        correto = "ÁREA VERDE"
                    cel[col + "_atual"] = atual
                    cel[col + "_correto"] = correto
                ws.append([dt.strftime("%Y-%m-%d"), dia_long, sgn, ra, nome,
                           cel["manha_atual"], cel["manha_correto"],
                           cel["tarde_atual"], cel["tarde_correto"],
                           cel["noite_atual"], cel["noite_correto"], None])
                r = ws.max_row
                for c in (7, 9, 11):
                    ws.cell(r, c).fill = AMARELO
    for col, w in zip("ABCDEFGHIJKL", (11, 11, 8, 12, 28, 20, 20, 20, 20, 18, 18, 24)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
