# -*- coding: utf-8 -*-
"""
publicar_lovable.py — Publicador determinístico de escalas no banco do Lovable (Supabase).

POR QUE EXISTE
--------------
As duas importações da aba "Visualizar Escalas" são incompletas e divergem entre si:
  • "Importar template em lote"  -> monta blocos, mas gera escala uniforme (sem área
    verde), deixa block_student_distribution.service_id NULL e a carga horária = 0.
  • "Importar Escala (formato diário)" -> escreve só escalas_diarias (com área verde),
    mas NÃO calcula carga horária nem monta o plano de blocos.

Resultado: a última importação corrige uma estrutura e deixa a outra defasada, e o
painel/Excel (que leem o "plano de blocos") divergem da escala real. Este módulo grava
TODAS as estruturas de forma consistente, num passo só, a partir da saída do gerador.

ESTRUTURAS GRAVADAS (de forma coerente, numa transação)
-------------------------------------------------------
  1. rodizios_escala            (datas, num_semanas, opcao_total_sg)
  2. specialty_block_config     (blocos com nomes)
  3. block_student_distribution (block_index -> serviço [services], horários)
  4. block_week_assignments     (subgrupo × semana -> bloco)  [fonte do painel/Excel]
  5. students.subgrupo          (divisão dos subgrupos do rodízio)
  6. semanas_rodizio            (semanas com data_inicio/fim)
  7. escalas_diarias            (serviço por período [servicos_escala] + CH calculada)

ENTRADAS (saída do gerador)
---------------------------
  • template_*.xlsx        (abas: _Blocos, Distribuição, _Serviços, Semana Padrão, _Subgrupos)
  • importar_lovable_*.xlsx (aba "Correções <cod>": 1 linha por aluno × dia × período)
  • definicoes_*.json      (metadados: especialidade, turma, grupo, data_inicio, num_semanas)

SEGREDO
-------
A connection string fica em SUPABASE_DB_URL (variável de ambiente / .env fora do git).
Nunca é versionada nem impressa.

USO
---
  # validar (não escreve nada — imprime o que faria):
  python publicar_lovable.py --template T.xlsx --importar I.xlsx --definicoes D.json \
      --codigo R1-GO-T6 --dry-run

  # publicar de verdade (precisa de SUPABASE_DB_URL no ambiente):
  python publicar_lovable.py --template T.xlsx --importar I.xlsx --definicoes D.json \
      --codigo R1-GO-T6 --opcao-sg 8
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any

import openpyxl

# ───────────────────────── helpers ─────────────────────────

DIA_LONG_TO_ABBR = {
    "segunda": "Seg", "terca": "Ter", "terça": "Ter", "quarta": "Qua",
    "quinta": "Qui", "sexta": "Sex", "sabado": "Sáb", "sábado": "Sáb", "domingo": "Dom",
}


def _norm(s: Any) -> str:
    """Normaliza para casamento: sem acento, minúsculo, espaços colapsados."""
    s = "".join(c for c in unicodedata.normalize("NFKD", str(s or "")) if not unicodedata.combining(c))
    s = "".join(c if c.isalnum() else " " for c in s.lower())
    return " ".join(s.split())


def _hours_between(hi: str | None, hf: str | None) -> float:
    """Duração em horas entre dois horários 'HH:MM'. 0 se inválido/ausente."""
    if not hi or not hf:
        return 0.0
    try:
        h1, m1 = (int(x) for x in str(hi).split(":")[:2])
        h2, m2 = (int(x) for x in str(hf).split(":")[:2])
    except (ValueError, AttributeError):
        return 0.0
    diff = (h2 * 60 + m2) - (h1 * 60 + m1)
    return round(diff / 60.0, 1) if diff > 0 else 0.0


def _is_free_or_theoretical(nome: str | None) -> bool:
    u = (nome or "").upper()
    return any(k in u for k in ("AREA VERDE", "ÁREA VERDE", "LIVRE", "AULA", "CORREÇ", "ENAMED", "SIMULAÇÃO", "BLOQUEIO"))


# ───────────────────────── catálogo de serviços ─────────────────────────


@dataclass
class ServiceInfo:
    """Um serviço resolvido contra o banco."""
    servicos_escala_id: str          # id em servicos_escala (usado por escalas_diarias)
    services_master_id: str | None   # id em services (usado pelas tabelas de bloco)
    nome: str
    him: str | None = None           # horário início manhã
    hfm: str | None = None
    hit: str | None = None           # horário início tarde
    hft: str | None = None
    is_area_verde: bool = False


class ServiceCatalog:
    """
    Resolve um nome de serviço (como aparece nos arquivos do gerador) para os IDs reais
    do banco. Mantém o casamento por nome normalizado, com suporte a apelidos.

    É alimentado por fetch_service_catalog() (que lê o banco) — ou injetado em testes.
    """

    def __init__(self, by_norm: dict[str, ServiceInfo], aliases: dict[str, str] | None = None):
        self._by_norm = by_norm
        self._alias = {_norm(k): _norm(v) for k, v in (aliases or {}).items()}

    def resolve(self, nome: str) -> ServiceInfo:
        key = _norm(nome)
        if key in self._by_norm:
            return self._by_norm[key]
        if key in self._alias and self._alias[key] in self._by_norm:
            return self._by_norm[self._alias[key]]
        raise KeyError(
            f"Serviço não encontrado no banco: {nome!r} (normalizado: {key!r}). "
            f"Ajuste o nome no arquivo ou a coluna 'nome do serviço cadastrado no lovable' "
            f"na aba _Serviços, ou adicione um apelido."
        )

    def has(self, nome: str) -> bool:
        key = _norm(nome)
        return key in self._by_norm or (key in self._alias and self._alias[key] in self._by_norm)


# Apelidos por especialidade: nomes que o GERADOR exporta -> nome cadastrado no Lovable.
# São aplicados SÓ dentro da especialidade certa (build_plan/fetch_service_catalog são
# por-especialidade), então é seguro mapear "PS SCA"->"PS GO SCA" sem afetar CIRURGIA/PED,
# que têm seus próprios "PS SCA"/"PNAR SCA".
_SPECIALTY_ALIASES: dict[str, dict[str, str]] = {
    "ginecologia e obstetricia": {
        "USG MANDIC": "USG GO MANDIC",
        "ENF SCA PUERPÉRIO": "ENFERMARIA PUERPERIO SCA",
        "ENF SCA": "ENFERMARIA PUERPERIO SCA",
        "ENFERMARIA SCA ALTO RISCO": "ENFERMARIA ALTO RISCO SCA",
        "ENF ALTO RISCO": "ENFERMARIA ALTO RISCO SCA",
        "PS SCA": "PS GO SCA",
        "PNAR SCA": "PNAR GO SCA",
        "PNAR MANDIC": "AMBULATÓRIO GO",
        "ambulatório mandic": "AMBULATÓRIO GO",
        "amb": "AMBULATÓRIO GO",
    },
}


def specialty_aliases(especialidade: str) -> dict[str, str]:
    """Apelidos padrão para a especialidade (casados por _esp_lookups)."""
    out: dict[str, str] = {}
    for esp in _esp_lookups(especialidade):
        out.update(_SPECIALTY_ALIASES.get(_norm(esp), {}))
    return out


def fetch_service_catalog(query_fn, especialidade: str, alias_extra: dict[str, str] | None = None) -> ServiceCatalog:
    """
    Lê do banco os serviços da especialidade (e ÁREA VERDE, que é 'TODAS') e monta o catálogo.

    `query_fn(sql) -> list[dict]` executa um SELECT e devolve linhas como dicts.
    Une `servicos_escala` (id + horários + link para o mestre) com `services` (id mestre).
    """
    esps = _esp_lookups(especialidade)
    esps_sql = ", ".join("'" + e.replace("'", "''") + "'" for e in esps)
    rows = query_fn(
        f"""
        SELECT se.id            AS se_id,
               se.nome          AS nome,
               se.service_id    AS master_id,
               se.horario_inicio_manha him, se.horario_fim_manha hfm,
               se.horario_inicio_tarde hit, se.horario_fim_tarde hft
        FROM servicos_escala se
        WHERE se.especialidade IN ({esps_sql})
           OR se.nome ILIKE '%AREA VERDE%' OR se.nome ILIKE '%ÁREA VERDE%'
        """
    )
    by_norm: dict[str, ServiceInfo] = {}
    for r in rows:
        info = ServiceInfo(
            servicos_escala_id=r["se_id"],
            services_master_id=r.get("master_id"),
            nome=r["nome"],
            him=r.get("him"), hfm=r.get("hfm"), hit=r.get("hit"), hft=r.get("hft"),
            is_area_verde=_norm(r["nome"]) == _norm("ÁREA VERDE"),
        )
        # primeira ocorrência vence (evita duplicatas de servicos_escala com mesmo nome)
        by_norm.setdefault(_norm(r["nome"]), info)
    return ServiceCatalog(by_norm, alias_extra)


def _esp_lookups(esp: str) -> list[str]:
    """Mesmos apelidos de especialidade que o app usa (espLookups)."""
    s = {esp}
    U = esp.upper()
    if "CIRURGIA" in U or "CIRÚRGICA" in U or "CIRURGICA" in U:
        s |= {"CLÍNICA CIRÚRGICA", "CIRURGIA"}
    if "GINECOLOGIA" in U:
        s |= {"GO", "GINECOLOGIA E OBSTETRÍCIA"}
    if "SAÚDE MENTAL" in U:
        s |= {"SM", "SAÚDE MENTAL"}
    if "PEDIATRIA" in U:
        s |= {"PED"}
    if "CLÍNICA MÉDICA" in U:
        s |= {"CM"}
    if "MFC" in U or "FAMÍLIA" in U:
        s |= {"MFC", "MEDICINA DE FAMÍLIA E COMUNIDADE"}
    return sorted(s)


# ───────────────────────── entradas (arquivos do gerador) ─────────────────────────


@dataclass
class Inputs:
    especialidade: str
    turma: str
    grupo: str
    data_inicio: date
    num_semanas: int
    # _Serviços: nome no arquivo (normalizado) -> nome cadastrado no lovable
    servico_alias: dict[str, str]
    # _Blocos: índice (0-based) -> nome do bloco
    blocos: list[str]
    blocos_semanas: list[int]
    # Distribuição: subgrupo (str) -> [nome_bloco por semana]
    distribuicao: dict[str, list[str]]
    # Semana Padrão: block_index -> serviço representativo (nome no arquivo)
    bloco_servico: dict[int, str]
    # _Subgrupos: subgrupo (str) -> [ra]
    subgrupos: dict[str, list[str]]
    # Correções: lista de dicts {ra, data, dia_semana, subgrupo, manha, tarde}
    diario: list[dict]


def load_inputs(template_path: str, importar_path: str, definicoes_path: str) -> Inputs:
    defs = json.load(open(definicoes_path, encoding="utf-8"))
    esp = defs["especialidade"]
    turma = defs["turma"]
    grupo = (defs.get("grupo") or "").replace("Grupo ", "").strip()
    data_inicio = datetime.strptime(defs["data_inicio"], "%Y-%m-%d").date()
    num_semanas = int(defs.get("num_semanas") or 8)

    tb = openpyxl.load_workbook(template_path, data_only=True)

    # _Serviços -> alias (Nome/Nome Curto -> nome cadastrado no lovable)
    servico_alias: dict[str, str] = {}
    ws = tb["_Serviços"]
    for row in list(ws.iter_rows(values_only=True))[1:]:
        nome, curto, lovable = (row + (None, None, None))[:3]
        if nome and lovable:
            servico_alias[_norm(nome)] = lovable
        if curto and lovable:
            servico_alias[_norm(curto)] = lovable

    # _Blocos
    blocos: list[str] = []
    blocos_semanas: list[int] = []
    for row in list(tb["_Blocos"].iter_rows(values_only=True))[1:]:
        idx, nome, semanas = (row + (None, None, None))[:3]
        if nome:
            blocos.append(str(nome))
            blocos_semanas.append(int(semanas or 1))

    # Distribuição
    distribuicao: dict[str, list[str]] = {}
    drows = list(tb["Distribuição"].iter_rows(values_only=True))
    for row in drows[1:]:
        sub = row[0]
        if sub is None:
            continue
        distribuicao[str(sub)] = [str(c) if c else "" for c in row[1:1 + num_semanas]]

    # Semana Padrão -> serviço representativo por bloco (serviço não-livre mais frequente)
    from collections import Counter
    bloco_serv_counts: dict[int, Counter] = {}
    sp = tb["Semana Padrão"]
    sp_hdr = [c.value for c in sp[1]]
    ci_bloco = sp_hdr.index("Bloco")
    ci_serv = sp_hdr.index("Serviço")
    for row in list(sp.iter_rows(values_only=True))[1:]:
        bi = row[ci_bloco]
        serv = row[ci_serv]
        if bi is None or not serv:
            continue
        bi0 = int(bi) - 1  # planilha é 1-based
        if _is_free_or_theoretical(serv):
            continue
        bloco_serv_counts.setdefault(bi0, Counter())[serv] += 1
    bloco_servico = {bi: cnt.most_common(1)[0][0] for bi, cnt in bloco_serv_counts.items()}

    # _Subgrupos -> RAs
    subgrupos: dict[str, list[str]] = {}
    for row in list(tb["_Subgrupos"].iter_rows(values_only=True))[1:]:
        sub = row[0]
        if sub is None:
            continue
        ras = []
        for cell in row[2:]:
            if cell and " - " in str(cell):
                ras.append(str(cell).split(" - ", 1)[0].strip())
        subgrupos[str(sub)] = ras

    # Correções (importar_lovable) -> diário
    ib = openpyxl.load_workbook(importar_path, data_only=True)
    corr_name = next((n for n in ib.sheetnames if n.lower().startswith("correç") or n.lower().startswith("correc")), None)
    if corr_name is None:
        raise ValueError("Aba 'Correções <cod>' não encontrada no importar_lovable.")
    cw = ib[corr_name]
    crows = list(cw.iter_rows(values_only=True))
    ch = {k: i for i, k in enumerate(crows[0])}
    diario: list[dict] = []
    for row in crows[1:]:
        ra = row[ch["ra"]]
        if not ra:
            continue
        diario.append({
            "ra": str(ra).strip(),
            "data": str(row[ch["data"]]).strip(),
            "dia_semana": str(row[ch["dia_semana"]]).strip(),
            "subgrupo": str(row[ch["subgrupo"]]).strip(),
            "manha": row[ch.get("manha_correto", -1)] if "manha_correto" in ch else None,
            "tarde": row[ch.get("tarde_correto", -1)] if "tarde_correto" in ch else None,
        })

    return Inputs(
        especialidade=esp, turma=turma, grupo=grupo, data_inicio=data_inicio,
        num_semanas=num_semanas, servico_alias=servico_alias, blocos=blocos,
        blocos_semanas=blocos_semanas, distribuicao=distribuicao,
        bloco_servico=bloco_servico, subgrupos=subgrupos, diario=diario,
    )


# ───────────────────────── construção do plano (lógica pura) ─────────────────────────


@dataclass
class Plan:
    """Tudo que será gravado, já resolvido para IDs. Determinístico a partir das entradas."""
    rodizio: dict
    config: dict                       # specialty_block_config
    block_dist: list[dict]             # block_student_distribution (block_index, service_master_id, horários)
    block_week: list[dict]             # block_week_assignments (subgrupo, semana, block_index, service_master_id, ra)
    students_subgrupo: list[dict]      # (ra, subgrupo)
    semanas: list[dict]                # semanas_rodizio
    diarias: list[dict]                # escalas_diarias (ra, data, dia, subgrupo, sm_id, st_id, ch_m, ch_t)
    warnings: list[str] = field(default_factory=list)


def build_plan(inp: Inputs, codigo: str, catalog: ServiceCatalog, opcao_total_sg: int | None) -> Plan:
    warnings: list[str] = []

    def alias(nome: str) -> str:
        return inp.servico_alias.get(_norm(nome), nome)

    def resolve(nome: str) -> ServiceInfo:
        return catalog.resolve(alias(nome))

    # ----- rodízio -----
    data_fim = inp.data_inicio + timedelta(days=inp.num_semanas * 7 - 1)
    rodizio = {
        "codigo": codigo,
        "especialidade": inp.especialidade,
        "turma": inp.turma,
        "data_inicio": inp.data_inicio.isoformat(),
        "data_fim": data_fim.isoformat(),
        "numero_semanas": inp.num_semanas,
        "opcao_total_sg": opcao_total_sg,
    }

    # ----- config de blocos -----
    config = {
        "especialidade": inp.especialidade,
        "turma": inp.turma,
        "block_type": f"custom_{len(inp.blocos)}",
        "blocks": [{"name": nome, "weeks": int(w)} for nome, w in zip(inp.blocos, inp.blocos_semanas)],
    }

    # ----- mapa nome_do_bloco -> block_index -----
    bloco_idx_by_name = {_norm(nome): i for i, nome in enumerate(inp.blocos)}

    # ----- block_student_distribution (serviço por bloco) -----
    block_dist: list[dict] = []
    for bi, nome in enumerate(inp.blocos):
        serv_nome = inp.bloco_servico.get(bi)
        if not serv_nome:
            warnings.append(f"Bloco {bi+1} ({nome}) sem serviço na Semana Padrão — service_id ficará NULL.")
            si = None
        else:
            si = resolve(serv_nome)
        block_dist.append({
            "block_index": bi,
            "service_master_id": si.services_master_id if si else None,
            "him": si.him if si else "07:00", "hfm": si.hfm if si else "12:00",
            "hit": si.hit if si else "13:00", "hft": si.hft if si else "17:00",
            "qtd_alunos": max((len(r) for r in inp.subgrupos.values()), default=3),
        })

    # ----- block_week_assignments (subgrupo × semana -> bloco) -----
    block_week: list[dict] = []
    for sub, blocos_semana in inp.distribuicao.items():
        ras = inp.subgrupos.get(sub, [])
        for w, nome_bloco in enumerate(blocos_semana, start=1):
            if not nome_bloco:
                continue
            bi = bloco_idx_by_name.get(_norm(nome_bloco))
            if bi is None:
                warnings.append(f"Distribuição SG{sub} Sem{w}: bloco {nome_bloco!r} não encontrado em _Blocos.")
                continue
            serv_master = block_dist[bi]["service_master_id"]
            for ra in ras:
                block_week.append({
                    "ra": ra, "subgrupo": sub, "semana": w,
                    "block_index": bi, "service_master_id": serv_master,
                })

    # ----- students.subgrupo -----
    students_subgrupo = [
        {"ra": ra, "subgrupo": sub}
        for sub, ras in inp.subgrupos.items() for ra in ras
    ]

    # ----- semanas_rodizio -----
    semanas = []
    for w in range(1, inp.num_semanas + 1):
        di = inp.data_inicio + timedelta(days=(w - 1) * 7)
        df = di + timedelta(days=6)
        semanas.append({"numero_semana": w, "data_inicio": di.isoformat(), "data_fim": df.isoformat()})

    # ----- escalas_diarias (serviço por período + CH) -----
    diarias: list[dict] = []
    for d in inp.diario:
        def per(nome):
            if not nome or _norm(nome) == _norm("VAZIO") or str(nome).strip() == "":
                return None, 0.0, 0.0
            info = resolve(nome)
            chm = 0.0 if info.is_area_verde else _hours_between(info.him, info.hfm)
            cht = 0.0 if info.is_area_verde else _hours_between(info.hit, info.hft)
            return info, chm, cht

        sm, sm_chm, _ = per(d["manha"])
        st, _, st_cht = per(d["tarde"])
        try:
            data_obj = datetime.strptime(d["data"], "%Y-%m-%d").date()
            semana = (data_obj - inp.data_inicio).days // 7 + 1
        except ValueError:
            semana = None
        diarias.append({
            "ra": d["ra"], "data": d["data"],
            "dia_semana": DIA_LONG_TO_ABBR.get(_norm(d["dia_semana"]), d["dia_semana"]),
            "subgrupo": d["subgrupo"], "semana": semana,
            "servico_manha_id": sm.servicos_escala_id if sm else None,
            "servico_tarde_id": st.servicos_escala_id if st else None,
            "ch_manha": sm_chm if sm else 0.0,
            "ch_tarde": st_cht if st else 0.0,
        })

    return Plan(
        rodizio=rodizio, config=config, block_dist=block_dist, block_week=block_week,
        students_subgrupo=students_subgrupo, semanas=semanas, diarias=diarias, warnings=warnings,
    )


# ───────────────────────── execução (transação) ─────────────────────────


def publish(plan: Plan, dsn: str, grupo: str) -> None:
    """Grava o plano no banco numa transação. Idempotente por rodízio (substitui)."""
    import psycopg2
    import psycopg2.extras

    conn = psycopg2.connect(dsn, sslmode="require")
    try:
        conn.autocommit = False
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        _execute_plan(cur, plan, grupo)
        conn.commit()
        print("✅ Publicado com sucesso (transação confirmada).")
    except Exception:
        conn.rollback()
        print("❌ Erro — rollback aplicado (nada foi gravado).", file=sys.stderr)
        raise
    finally:
        conn.close()


def _execute_plan(cur, plan: Plan, grupo: str) -> None:
    r = plan.rodizio
    # 1) rodízio (buscar por código -> atualizar; senão inserir).
    #    Não há UNIQUE em codigo, então não usamos ON CONFLICT.
    cur.execute("SELECT id FROM rodizios_escala WHERE codigo=%s", (r["codigo"],))
    row = cur.fetchone()
    if row:
        rodizio_id = row["id"]
        cur.execute(
            """UPDATE rodizios_escala SET data_inicio=%(data_inicio)s, data_fim=%(data_fim)s,
               numero_semanas=%(numero_semanas)s, opcao_total_sg=%(opcao_total_sg)s WHERE id=%(id)s""",
            {**r, "id": rodizio_id})
    else:
        cur.execute(
            """INSERT INTO rodizios_escala (codigo, especialidade, turma, data_inicio, data_fim, numero_semanas, opcao_total_sg)
               VALUES (%(codigo)s,%(especialidade)s,%(turma)s,%(data_inicio)s,%(data_fim)s,%(numero_semanas)s,%(opcao_total_sg)s)
               RETURNING id""", r)
        rodizio_id = cur.fetchone()["id"]

    # 2) config de blocos (buscar por especialidade+turma -> atualizar; senão inserir).
    blocks_json = json.dumps(plan.config["blocks"], ensure_ascii=False)
    cur.execute("SELECT id FROM specialty_block_config WHERE especialidade=%s AND turma=%s",
                (plan.config["especialidade"], plan.config["turma"]))
    row = cur.fetchone()
    if row:
        config_id = row["id"]
        cur.execute("UPDATE specialty_block_config SET block_type=%s, blocks=%s, updated_at=now() WHERE id=%s",
                    (plan.config["block_type"], blocks_json, config_id))
    else:
        cur.execute(
            """INSERT INTO specialty_block_config (especialidade, turma, block_type, blocks)
               VALUES (%s,%s,%s,%s) RETURNING id""",
            (plan.config["especialidade"], plan.config["turma"], plan.config["block_type"], blocks_json))
        config_id = cur.fetchone()["id"]

    # 3) block_student_distribution (limpa e reinsere) -> guarda dist_id por block_index
    cur.execute("DELETE FROM block_student_distribution WHERE config_id=%s", (config_id,))
    dist_id_by_idx: dict[int, str] = {}
    for bd in plan.block_dist:
        cur.execute(
            """
            INSERT INTO block_student_distribution
              (config_id, block_index, service_id, qtd_alunos, periodo_tipo,
               horario_inicio, horario_fim, horario_inicio_tarde, horario_fim_tarde)
            VALUES (%s,%s,%s,%s,'manha_tarde',%s,%s,%s,%s)
            RETURNING id
            """,
            (config_id, bd["block_index"], bd["service_master_id"], bd["qtd_alunos"],
             bd["him"], bd["hfm"], bd["hit"], bd["hft"]))
        dist_id_by_idx[bd["block_index"]] = cur.fetchone()["id"]

    # mapa ra -> student_id (só dos alunos deste rodízio)
    ras = sorted({s["ra"] for s in plan.students_subgrupo})
    cur.execute("SELECT id, ra FROM students WHERE ra = ANY(%s)", (ras,))
    sid_by_ra = {row["ra"]: row["id"] for row in cur.fetchall()}
    faltando = [ra for ra in ras if ra not in sid_by_ra]
    if faltando:
        raise RuntimeError(f"Alunos não encontrados em students: {faltando}")

    # 4) block_week_assignments (limpa as dos blocos deste config e reinsere)
    cur.execute(
        "DELETE FROM block_week_assignments WHERE distribution_id IN "
        "(SELECT id FROM block_student_distribution WHERE config_id=%s)", (config_id,))
    for bw in plan.block_week:
        sid = sid_by_ra.get(bw["ra"])
        if not sid:
            continue
        cur.execute(
            "INSERT INTO block_week_assignments (distribution_id, semana, student_id, service_id) VALUES (%s,%s,%s,%s)",
            (dist_id_by_idx[bw["block_index"]], bw["semana"], sid, bw["service_master_id"]))

    # 5) students.subgrupo
    for s in plan.students_subgrupo:
        cur.execute(
            "UPDATE students SET subgrupo=%s, updated_at=now() WHERE ra=%s AND turma=%s",
            (s["subgrupo"], s["ra"], r["turma"]))

    # 6) semanas_rodizio (limpa e reinsere) -> semana_id por numero
    cur.execute("DELETE FROM semanas_rodizio WHERE rodizio_id=%s", (rodizio_id,))
    semana_id_by_num: dict[int, str] = {}
    for sm in plan.semanas:
        cur.execute(
            "INSERT INTO semanas_rodizio (rodizio_id, numero_semana, data_inicio, data_fim) VALUES (%s,%s,%s,%s) RETURNING id",
            (rodizio_id, sm["numero_semana"], sm["data_inicio"], sm["data_fim"]))
        semana_id_by_num[sm["numero_semana"]] = cur.fetchone()["id"]

    # 7) escalas_diarias (limpa e reinsere)
    cur.execute("DELETE FROM escalas_diarias WHERE rodizio_id=%s", (rodizio_id,))
    for e in plan.diarias:
        cur.execute(
            """
            INSERT INTO escalas_diarias
              (ra, rodizio_id, semana_id, data, dia_semana, subgrupo, grupo,
               servico_manha_id, servico_tarde_id, ch_manha, ch_tarde)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (e["ra"], rodizio_id, semana_id_by_num.get(e["semana"]), e["data"], e["dia_semana"],
             e["subgrupo"], grupo, e["servico_manha_id"], e["servico_tarde_id"],
             e["ch_manha"], e["ch_tarde"]))


# ───────────────────────── dry-run / resumo ─────────────────────────


def print_summary(plan: Plan) -> None:
    ch_total = sum((d["ch_manha"] or 0) + (d["ch_tarde"] or 0) for d in plan.diarias)
    n_alunos = len({s["ra"] for s in plan.students_subgrupo})
    print("── RESUMO DO PLANO (dry-run, nada gravado) ─────────────────────")
    print(f"  Rodízio        : {plan.rodizio['codigo']}  ({plan.rodizio['especialidade']} / {plan.rodizio['turma']})")
    print(f"  Período        : {plan.rodizio['data_inicio']} → {plan.rodizio['data_fim']}  ({plan.rodizio['numero_semanas']} sem)")
    print(f"  Config blocos  : {plan.config['block_type']}  ({len(plan.config['blocks'])} blocos)")
    print(f"  Alunos         : {n_alunos}   Subgrupos: {len(plan.students_subgrupo and set(s['subgrupo'] for s in plan.students_subgrupo))}")
    print(f"  block_dist     : {len(plan.block_dist)} linhas")
    print(f"  block_week     : {len(plan.block_week)} linhas")
    print(f"  semanas        : {len(plan.semanas)}")
    print(f"  escalas_diarias: {len(plan.diarias)} linhas")
    print(f"  CH prática     : {ch_total:.0f}h  (média {ch_total / n_alunos:.0f}h/aluno)" if n_alunos else "")
    print("  block_student_distribution (block_index -> serviço):")
    for bd in plan.block_dist:
        print(f"     [{bd['block_index']}] master={bd['service_master_id']}  {bd['him']}-{bd['hfm']} / {bd['hit']}-{bd['hft']}")
    if plan.warnings:
        print("  ⚠️  AVISOS:")
        for w in plan.warnings:
            print(f"     - {w}")
    print("────────────────────────────────────────────────────────────────")


# ───────────────────────── CLI ─────────────────────────


def main(argv=None):
    ap = argparse.ArgumentParser(description="Publica uma escala completa no banco do Lovable (Supabase).")
    ap.add_argument("--template", required=True, help="template_*.xlsx (gerador)")
    ap.add_argument("--importar", required=True, help="importar_lovable_*.xlsx (gerador)")
    ap.add_argument("--definicoes", required=True, help="definicoes_*.json (gerador)")
    ap.add_argument("--codigo", required=True, help="código do rodízio (ex: R1-GO-T6)")
    ap.add_argument("--opcao-sg", type=int, default=None, help="opcao_total_sg do rodízio (ex: 8)")
    ap.add_argument("--dry-run", action="store_true", help="não grava — só mostra o resumo do plano")
    args = ap.parse_args(argv)

    inp = load_inputs(args.template, args.importar, args.definicoes)

    if args.dry_run:
        print("DRY-RUN: catálogo de serviços não consultado (use a publicação real para resolver IDs).")
        print(f"  Entradas OK: {len(inp.blocos)} blocos, {len(inp.subgrupos)} subgrupos, {len(inp.diario)} linhas diárias.")
        return 0

    dsn = os.environ.get("SUPABASE_DB_URL")
    if not dsn:
        print("ERRO: defina SUPABASE_DB_URL no ambiente (.env). Veja .env.example.", file=sys.stderr)
        return 2

    import psycopg2
    import psycopg2.extras
    conn = psycopg2.connect(dsn, sslmode="require")
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        def query_fn(sql):
            cur.execute(sql)
            return cur.fetchall()

        catalog = fetch_service_catalog(query_fn, inp.especialidade, specialty_aliases(inp.especialidade))
        plan = build_plan(inp, args.codigo, catalog, args.opcao_sg)
        print_summary(plan)
        _execute_plan(cur, plan, inp.grupo)
        conn.commit()
        print("✅ Publicado com sucesso.")
    except Exception:
        conn.rollback()
        print("❌ Erro — rollback (nada gravado).", file=sys.stderr)
        raise
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
